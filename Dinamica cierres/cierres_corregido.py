import pandas as pd
import mysql.connector
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
from datetime import datetime

## ================== CONFIGURACIÓN ==================
DB_CONFIG = {
    'host': '10.108.34.32',
    'port': 33063,
    'user': 'ccot',
    'password': 'ccot',
    'database': 'ccot',
    'charset': 'utf8mb4',
    'use_unicode': True
}

SQL_QUERY = """
SELECT *
FROM back_informe
WHERE CUENTA <> '0'
  AND ESTADO = 'COMPLETADO';
"""

OUTPUT_XLSX = "Informe_Cierres_Diario_Dinamico.xlsx"
COL_TIPO_TRABAJO = "TRABAJO_WF"
COL_FECHA = "FECHA_AGENDA"
COL_ALIADO = "ALIADO"
COL_AREA = "AREA"
COL_TIPO_CLIENTE = "TIPO_CLIENTE"
## ====================================================

# COLORES DEL DISEÑO ESTANDARIZADO - NEGRO Y ROJO
COLOR_HEADER_BLACK = "000000"           # Negro para encabezados principales
COLOR_HEADER_RED = "C00000"             # Rojo oscuro para encabezados secundarios
COLOR_LIGHT_RED = "F4B084"              # Rojo claro para filas de datos
COLOR_DARK_RED = "C00000"               # Rojo oscuro para énfasis
COLOR_WHITE = "FFFFFF"                  # Blanco para texto sobre negro
COLOR_LIGHT_GRAY = "D9D9D9"             # Gris claro para alternancia
COLOR_MENU_BG = "1F1F1F"                # Negro suave para fondo de menú
COLOR_CARD_BG = "2D2D2D"                # Gris oscuro para tarjetas
COLOR_BORDER = "808080"                 # Gris para bordes


def fetch_data():
    """Ejecuta la consulta SQL y devuelve un DataFrame"""
    print(f"🔌 Conectando a la base de datos...")
    
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        df = pd.read_sql(SQL_QUERY, conn)
        conn.close()
        print(f"✅ Datos obtenidos: {len(df)} registros")
        return df
        
    except mysql.connector.Error as e:
        print(f"❌ Error de conexión: {e}")
        raise
    except Exception as e:
        print(f"❌ Error inesperado: {e}")
        raise


def build_daily_pivot(df: pd.DataFrame):
    """Crea una tabla dinámica por día"""
    
    required_cols = [COL_TIPO_TRABAJO, COL_FECHA]
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"❌ Columna '{col}' no encontrada en los datos")

    df[COL_FECHA] = pd.to_datetime(df[COL_FECHA], errors="coerce")
    df = df.dropna(subset=[COL_FECHA])
    
    df['FECHA_FORMATEADA'] = df[COL_FECHA].dt.strftime('%Y-%m-%d')
    df['FECHA_DATE'] = df[COL_FECHA].dt.date
    
    try:
        pivot_main = pd.pivot_table(
            df,
            values='FECHA_FORMATEADA',
            index=COL_TIPO_TRABAJO,
            columns='FECHA_DATE',
            aggfunc='count',
            fill_value=0
        )
        
        pivot_main['Total general'] = pivot_main.sum(axis=1)
        totals_row = pivot_main.sum(axis=0).to_frame().T
        totals_row.index = ['Total general']
        pivot_main = pd.concat([pivot_main, totals_row])
        
    except Exception as e:
        print(f"⚠️ Error creando pivot principal: {e}")
        pivot_data = df.groupby([COL_TIPO_TRABAJO, 'FECHA_DATE']).size().unstack(fill_value=0)
        pivot_main = pivot_data.copy()
        pivot_main['Total general'] = pivot_main.sum(axis=1)
        totals_row = pivot_main.sum(axis=0).to_frame().T
        totals_row.index = ['Total general']
        pivot_main = pd.concat([pivot_main, totals_row])
    
    return pivot_main, df


def create_menu_sheet(wb, stats):
    """Crea una hoja de menú gerencial con diseño negro y rojo"""
    
    if 'MENÚ GERENCIAL' in wb.sheetnames:
        ws = wb['MENÚ GERENCIAL']
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet('MENÚ GERENCIAL', 0)
    
    ws.sheet_view.showGridLines = False
    
    # Configurar anchos de columna
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 2
    
    # ENCABEZADO PRINCIPAL - Negro con texto blanco
    ws.merge_cells('A1:D1')
    ws['A1'] = "📊 DASHBOARD GERENCIAL - CIERRES DIARIOS"
    ws['A1'].font = Font(bold=True, size=18, color=COLOR_WHITE)
    ws['A1'].fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Fecha de actualización - Rojo
    ws.merge_cells('A2:D2')
    ws['A2'] = f"Última actualización: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws['A2'].font = Font(size=11, color=COLOR_DARK_RED, italic=True, bold=True)
    ws['A2'].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 25
    
    # Espacio
    ws.row_dimensions[3].height = 10
    
    # ESTADÍSTICAS GENERALES
    border_card = Border(
        left=Side(style='medium', color=COLOR_BORDER),
        right=Side(style='medium', color=COLOR_BORDER),
        top=Side(style='medium', color=COLOR_BORDER),
        bottom=Side(style='medium', color=COLOR_BORDER)
    )
    
    current_row = 4
    
    # Card: Total de Registros
    ws.merge_cells(f'B{current_row}:C{current_row}')
    ws[f'B{current_row}'] = "RESUMEN GENERAL"
    ws[f'B{current_row}'].font = Font(bold=True, size=14, color=COLOR_WHITE)
    ws[f'B{current_row}'].fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
    ws[f'B{current_row}'].alignment = Alignment(horizontal="center")
    ws[f'B{current_row}'].border = border_card
    current_row += 1
    
    ws.merge_cells(f'B{current_row}:C{current_row}')
    ws[f'B{current_row}'] = f"Total de Cierres: {stats['total_registros']}"
    ws[f'B{current_row}'].font = Font(size=24, bold=True, color=COLOR_DARK_RED)
    ws[f'B{current_row}'].fill = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")
    ws[f'B{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
    ws[f'B{current_row}'].border = border_card
    ws.row_dimensions[current_row].height = 50
    
    current_row += 1
    ws.row_dimensions[current_row].height = 15
    current_row += 1
    
    # TARJETAS DE NAVEGACIÓN
    ws.merge_cells(f'B{current_row}:C{current_row}')
    ws[f'B{current_row}'] = "ACCESOS RÁPIDOS A REPORTES"
    ws[f'B{current_row}'].font = Font(bold=True, size=14, color=COLOR_WHITE)
    ws[f'B{current_row}'].fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
    ws[f'B{current_row}'].alignment = Alignment(horizontal="center")
    ws[f'B{current_row}'].border = border_card
    current_row += 1
    ws.row_dimensions[current_row].height = 10
    current_row += 1
    
    # Definir tarjetas de navegación
    cards = [
        {
            'title': '📅 VISIÓN DIARIA',
            'desc': 'Análisis por Tipo de Trabajo y Fecha',
            'sheet': 'Visión Diaria',
            'icon': '📊'
        },
        {
            'title': '🤝 RESUMEN POR ALIADO',
            'desc': 'Agrupación de cierres por Aliado',
            'sheet': 'Resumen por Aliado',
            'icon': '👥'
        },
        {
            'title': '📍 RESUMEN POR ÁREA',
            'desc': 'Distribución geográfica de cierres',
            'sheet': 'Resumen por Area',
            'icon': '🗺️'
        },
        {
            'title': '👤 RESUMEN POR TIPO CLIENTE',
            'desc': 'Segmentación por tipo de cliente',
            'sheet': 'Resumen por Tipo Cliente',
            'icon': '📋'
        }
    ]
    
    # Crear tarjetas en formato 2x2
    for idx, card in enumerate(cards):
        col_offset = (idx % 2)
        base_col = 'B' if col_offset == 0 else 'C'
        
        # Título de la tarjeta - Negro con texto blanco
        cell_ref = f'{base_col}{current_row}'
        ws[cell_ref] = f"{card['icon']} {card['title']}"
        ws[cell_ref].font = Font(bold=True, size=12, color=COLOR_WHITE)
        ws[cell_ref].fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
        ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center")
        ws[cell_ref].border = border_card
        
        # Descripción - Blanco con texto negro
        cell_ref_desc = f'{base_col}{current_row + 1}'
        ws[cell_ref_desc] = card['desc']
        ws[cell_ref_desc].font = Font(size=10, color="000000")
        ws[cell_ref_desc].fill = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")
        ws[cell_ref_desc].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[cell_ref_desc].border = border_card
        
        # Botón de acceso - Rojo con texto blanco
        cell_ref_btn = f'{base_col}{current_row + 2}'
        ws[cell_ref_btn] = "▶ VER REPORTE"
        ws[cell_ref_btn].font = Font(bold=True, size=11, color=COLOR_WHITE, underline="single")
        ws[cell_ref_btn].fill = PatternFill(start_color=COLOR_DARK_RED, end_color=COLOR_DARK_RED, fill_type="solid")
        ws[cell_ref_btn].alignment = Alignment(horizontal="center", vertical="center")
        ws[cell_ref_btn].border = border_card
        ws[cell_ref_btn].hyperlink = f"#'{card['sheet']}'!A1"
        
        # Ajustar altura de filas
        ws.row_dimensions[current_row].height = 30
        ws.row_dimensions[current_row + 1].height = 35
        ws.row_dimensions[current_row + 2].height = 30
        
        # Si completamos 2 columnas, avanzar a la siguiente fila
        if (idx + 1) % 2 == 0:
            current_row += 4
    
    # Si quedó una tarjeta impar, avanzar
    if len(cards) % 2 != 0:
        current_row += 4
    
    # INFORMACIÓN ADICIONAL
    current_row += 2
    ws.merge_cells(f'B{current_row}:C{current_row}')
    ws[f'B{current_row}'] = "ℹ️ INFORMACIÓN"
    ws[f'B{current_row}'].font = Font(bold=True, size=11, color=COLOR_WHITE)
    ws[f'B{current_row}'].fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
    ws[f'B{current_row}'].alignment = Alignment(horizontal="center")
    ws[f'B{current_row}'].border = border_card
    
    current_row += 1
    ws.merge_cells(f'B{current_row}:C{current_row}')
    info_text = (
        "• Haga clic en cualquier número de los reportes para ver el detalle\n"
        "• Las hojas de detalle se mostrarán automáticamente\n"
        "• Use los filtros en cada hoja para análisis específicos\n"
        "• Use el botón '⬅ MENÚ' para volver al inicio"
    )
    ws[f'B{current_row}'] = info_text
    ws[f'B{current_row}'].font = Font(size=9, color="000000")
    ws[f'B{current_row}'].fill = PatternFill(start_color=COLOR_LIGHT_GRAY, end_color=COLOR_LIGHT_GRAY, fill_type="solid")
    ws[f'B{current_row}'].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws[f'B{current_row}'].border = border_card
    ws.row_dimensions[current_row].height = 60
    
    print("✅ Hoja 'MENÚ GERENCIAL' creada con diseño negro y rojo")


def create_detail_sheet(wb, df, filter_dict, sheet_name):
    """Crea una hoja con datos filtrados y la deja VISIBLE (no oculta)"""
    
    # Extraer df si está en filter_dict
    if 'df' in filter_dict:
        df = filter_dict.pop('df')
    
    df_filtered = df.copy()
    for col, value in filter_dict.items():
        if col in df_filtered.columns and value is not None:
            if col == COL_FECHA:
                df_filtered = df_filtered[df_filtered[col].dt.date == value]
            else:
                df_filtered = df_filtered[df_filtered[col] == value]
    
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
    else:
        ws = wb.create_sheet(sheet_name)
    
    # NO OCULTAR LA HOJA - Comentamos esta línea
    # ws.sheet_state = 'hidden'
    ws.sheet_state = 'visible'  # Aseguramos que sea visible
    
    ws.sheet_view.showGridLines = False
    
    border = Border(left=Side(style='thin', color=COLOR_BORDER), 
                   right=Side(style='thin', color=COLOR_BORDER),
                   top=Side(style='thin', color=COLOR_BORDER), 
                   bottom=Side(style='thin', color=COLOR_BORDER))
    
    # Título con diseño negro y rojo
    ws['A1'] = f"Detalle Filtrado - {len(df_filtered)} registros"
    ws['A1'].font = Font(bold=True, size=14, color=COLOR_WHITE)
    ws['A1'].fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
    ws['A1'].border = border
    ws.row_dimensions[1].height = 30
    
    # Botón para volver al menú - Rojo
    ws['A2'] = "⬅ Volver al Menú"
    ws['A2'].font = Font(bold=True, size=11, color=COLOR_WHITE, underline="single")
    ws['A2'].fill = PatternFill(start_color=COLOR_DARK_RED, end_color=COLOR_DARK_RED, fill_type="solid")
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A2'].hyperlink = "#'MENÚ GERENCIAL'!A1"
    ws['A2'].border = border
    ws.row_dimensions[2].height = 25
    
    # Mostrar filtros aplicados
    row = 3
    for col, value in filter_dict.items():
        if value is not None:
            ws[f'A{row}'] = f"{col}: {value}"
            ws[f'A{row}'].font = Font(bold=True, size=10, color=COLOR_DARK_RED)
            ws[f'A{row}'].fill = PatternFill(start_color=COLOR_LIGHT_GRAY, end_color=COLOR_LIGHT_GRAY, fill_type="solid")
            ws[f'A{row}'].border = border
            row += 1
    
    start_row = row + 1
    
    # Encabezados de columnas - Negro con texto blanco
    for col_idx, col_name in enumerate(df_filtered.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
        cell.font = Font(bold=True, color=COLOR_WHITE)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Datos - Rojo claro alternado
    for row_idx, row_data in enumerate(df_filtered.itertuples(index=False), start_row + 1):
        # Alternar colores cada fila
        bg_color = COLOR_LIGHT_RED if row_idx % 2 == 0 else COLOR_WHITE
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Ajustar anchos de columna
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Agregar filtros automáticos
    try:
        ws.auto_filter.ref = f"A{start_row}:{ws.cell(row=start_row, column=len(df_filtered.columns)).column_letter}{start_row + len(df_filtered)}"
    except:
        pass
    
    print(f"✅ Hoja de detalle '{sheet_name}' creada (VISIBLE)")
    return sheet_name


def create_summary_sheet(wb, df, col_name, sheet_title, sheet_name_short):
    """Función genérica para crear hojas de resumen con diseño negro y rojo"""
    
    if col_name not in df.columns:
        print(f"⚠️ Columna {col_name} no encontrada, saltando hoja de resumen")
        return
    
    if sheet_title in wb.sheetnames:
        ws = wb[sheet_title]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_title)
    
    ws.sheet_view.showGridLines = False
    
    border = Border(left=Side(style='thin', color=COLOR_BORDER), 
                   right=Side(style='thin', color=COLOR_BORDER),
                   top=Side(style='thin', color=COLOR_BORDER), 
                   bottom=Side(style='thin', color=COLOR_BORDER))
    
    # Botón volver al menú - Rojo
    ws['A1'] = "⬅ MENÚ"
    ws['A1'].font = Font(bold=True, size=10, color=COLOR_WHITE, underline="single")
    ws['A1'].fill = PatternFill(start_color=COLOR_DARK_RED, end_color=COLOR_DARK_RED, fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].hyperlink = "#'MENÚ GERENCIAL'!A1"
    ws['A1'].border = border
    ws.row_dimensions[1].height = 25
    
    # Fecha de actualización - Negro
    ws['C1'] = "Fecha Actualización"
    ws['C1'].fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
    ws['C1'].font = Font(bold=True, color=COLOR_WHITE)
    ws['C1'].border = border
    ws['C1'].alignment = Alignment(horizontal="center", vertical="center")
    
    ws['D1'] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws['D1'].fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
    ws['D1'].font = Font(color=COLOR_WHITE, bold=True)
    ws['D1'].border = border
    ws['D1'].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.row_dimensions[2].height = 5
    
    # Título - Negro con texto blanco
    ws['A3'] = sheet_title
    ws['A3'].font = Font(bold=True, size=14, color=COLOR_DARK_RED)
    
    ws.row_dimensions[4].height = 5
    
    df_work = df.copy()
    
    if 'FECHA_DATE' not in df_work.columns:
        df_work['FECHA_DATE'] = df_work[COL_FECHA].dt.date
    
    # Manejar valores nulos y agrupaciones especiales para ALIADO
    if col_name == COL_ALIADO:
        aliado_wf_col = 'ALIADO_WF' if 'ALIADO_WF' in df_work.columns else COL_ALIADO
        red_col = 'RED' if 'RED' in df_work.columns else None
        
        df_work['ALIADO_AGRUPADO'] = df_work[COL_ALIADO].fillna('')
        mask_vacio = (df_work['ALIADO_AGRUPADO'] == '') | (df_work['ALIADO_AGRUPADO'].isna())
        df_work.loc[mask_vacio, 'ALIADO_AGRUPADO'] = df_work.loc[mask_vacio, aliado_wf_col]
        
        tabasco_aliados = [
            'TABASCO OC,LLC. SUCURSAL COLOMBIA',
            'TABASCO OC, LLC. SUCURSAL COLOMBIA',
            'CICSA COLOMBIA S.A.'
        ]
        mask_tabasco = (
            ((df_work[COL_ALIADO] == 'TABASCO') | (df_work[COL_ALIADO].isna()) | (df_work[COL_ALIADO] == '')) & 
            (df_work[aliado_wf_col].isin(tabasco_aliados))
        )
        df_work.loc[mask_tabasco, 'ALIADO_AGRUPADO'] = 'TABASCO'
        
        mask_conectar = (
            (df_work[COL_ALIADO].isna() | (df_work[COL_ALIADO] == '') | 
             df_work[COL_ALIADO].str.contains('CONECTAR', na=False) | 
             df_work[aliado_wf_col].str.contains('CONECTAR', na=False))
        )
        df_work.loc[mask_conectar, COL_ALIADO] = 'CONECTAR'
        df_work.loc[mask_conectar, 'ALIADO_AGRUPADO'] = 'CONECTAR'
        
        distribuidor_dth_list = [
            'ASECONES','ATENTO S.A','CINCO S.A.S.','CLARO COLOMBIA',
            'COMCEL ALL STAR COMUNICACIONES','COMCEL ALZATE ECHEVERRI Y CIA LTDA',
            'COMCEL AMAZONIA PUNTO COM LTDA','COMCEL ANCLU S.A.S',
            'COMCEL ARTECOM COMUNICACIONES S.A.S','COMCEL BEST COMUNICACIONES',
            'COMCEL BLACKMOVIL','COMCEL COLMOVILES COMUNICACIONES',
            'COMCEL DAVIDMOVIL S.A.S','COMCEL DIESTRO S.A.S',
            'COMCEL ISAMOVIL SAS','COMCEL MOBILE CENTER SAS',
            'COMCEL SMART PHONE LTDA.','COMCEL WAP COMUNICACIONES S.A.S',
            'DICO TELECOMUNICACIONES S.A.','FENIX GROUP COMUNICACIONES S.A.S',
            'JESMAR HURTADO Y COMPANIA S. EN C.','PHONEMOVIL LTDA',
            'REDESTELCO SAS','SOLUCIONES MOVILESRIO S.A.S.','TELCOS INGENIERIA S.A.','ZENIX SAS'
        ]
        
        if red_col:
            mask_distribuidor = (df_work[aliado_wf_col].isin(distribuidor_dth_list)) & (df_work[red_col] == 'DTH')
            df_work.loc[mask_distribuidor, 'ALIADO_AGRUPADO'] = 'DISTRIBUIDOR DTH'
        
        index_col = 'ALIADO_AGRUPADO'
    else:
        df_work[col_name] = df_work[col_name].fillna(f'SIN {col_name}')
        index_col = col_name
    
    try:
        pivot = pd.pivot_table(
            df_work,
            values='FECHA_FORMATEADA',
            index=index_col,
            columns='FECHA_DATE',
            aggfunc='count',
            fill_value=0
        )
        
        pivot['Total general'] = pivot.sum(axis=1)
        totals_row = pivot.sum(axis=0).to_frame().T
        totals_row.index = ['Total general']
        pivot = pd.concat([pivot, totals_row])
        
    except Exception as e:
        print(f"⚠️ Error creando pivot: {e}")
        return
    
    current_row = 5
    
    # Encabezado de columna principal - Negro
    ws.cell(row=current_row, column=1, value=col_name.title())
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
    ws.cell(row=current_row, column=1).font = Font(bold=True, color=COLOR_WHITE)
    ws.cell(row=current_row, column=1).border = border
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    
    date_columns = [col for col in pivot.columns if col != 'Total general']
    try:
        date_columns_sorted = sorted(date_columns)
    except:
        date_columns_sorted = date_columns
    
    # Encabezados de fechas - Negro
    col_idx = 2
    for date_col in date_columns_sorted:
        if hasattr(date_col, 'strftime'):
            date_str = date_col.strftime("%d/%m/%Y")
        else:
            date_str = str(date_col)
        
        ws.cell(row=current_row, column=col_idx, value=date_str)
        ws.cell(row=current_row, column=col_idx).fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
        ws.cell(row=current_row, column=col_idx).font = Font(bold=True, color=COLOR_WHITE)
        ws.cell(row=current_row, column=col_idx).border = border
        ws.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
        col_idx += 1
    
    # Encabezado Total - Negro
    ws.cell(row=current_row, column=col_idx, value="Total general")
    ws.cell(row=current_row, column=col_idx).fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
    ws.cell(row=current_row, column=col_idx).font = Font(bold=True, color=COLOR_WHITE)
    ws.cell(row=current_row, column=col_idx).border = border
    ws.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
    total_col_idx = col_idx
    
    current_row += 1
    
    # Filas de datos - Rojo claro
    for item in pivot.index:
        if item != 'Total general':
            ws.cell(row=current_row, column=1, value=item)
            ws.cell(row=current_row, column=1).fill = PatternFill(start_color=COLOR_LIGHT_RED, end_color=COLOR_LIGHT_RED, fill_type="solid")
            ws.cell(row=current_row, column=1).border = border
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            
            col_idx = 2
            for date_col in date_columns_sorted:
                value = pivot.loc[item, date_col]
                cell = ws.cell(row=current_row, column=col_idx, value=int(value) if value > 0 else "")
                cell.fill = PatternFill(start_color=COLOR_LIGHT_RED, end_color=COLOR_LIGHT_RED, fill_type="solid")
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Crear hipervínculo si hay datos
                if value > 0:
                    sheet_name = f"Det_{sheet_name_short}_{str(item)[:8]}_{date_col.strftime('%d%m') if hasattr(date_col, 'strftime') else str(date_col)[:10]}"
                    sheet_name = re.sub(r"[:\\/?*\[\]]", "_", sheet_name)[:31]
                    cell.hyperlink = f"#'{sheet_name}'!A1"
                    cell.font = Font(color=COLOR_DARK_RED, underline="single", bold=True)
                    
                    # Guardar información para crear hoja de detalle
                    if not hasattr(wb, '_detail_sheets'):
                        wb._detail_sheets = []
                    
                    filters = {index_col: item, COL_FECHA: date_col}
                    if col_name == COL_ALIADO:
                        filters['df'] = df_work
                    
                    wb._detail_sheets.append({
                        'sheet_name': sheet_name,
                        'filters': filters
                    })
                
                col_idx += 1
            
            # Total por fila
            total_value = pivot.loc[item, 'Total general']
            cell = ws.cell(row=current_row, column=total_col_idx, value=int(total_value) if total_value > 0 else "")
            cell.fill = PatternFill(start_color=COLOR_LIGHT_RED, end_color=COLOR_LIGHT_RED, fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            if total_value > 0:
                sheet_name = f"Det_{sheet_name_short}_{str(item)[:13]}_Total"
                sheet_name = re.sub(r"[:\\/?*\[\]]", "_", sheet_name)[:31]
                cell.hyperlink = f"#'{sheet_name}'!A1"
                cell.font = Font(color=COLOR_DARK_RED, underline="single", bold=True)
                
                if not hasattr(wb, '_detail_sheets'):
                    wb._detail_sheets = []
                
                filters = {index_col: item}
                if col_name == COL_ALIADO:
                    filters['df'] = df_work
                
                wb._detail_sheets.append({
                    'sheet_name': sheet_name,
                    'filters': filters
                })
            
            current_row += 1
    
    # Fila de totales - Negro con texto blanco
    total_row = current_row
    ws.cell(row=total_row, column=1, value="Total general")
    ws.cell(row=total_row, column=1).fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
    ws.cell(row=total_row, column=1).font = Font(bold=True, color=COLOR_WHITE)
    ws.cell(row=total_row, column=1).border = border
    
    col_idx = 2
    for date_col in date_columns_sorted:
        total_value = pivot.loc['Total general', date_col]
        cell = ws.cell(row=total_row, column=col_idx, value=int(total_value) if total_value > 0 else "")
        cell.fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
        cell.font = Font(bold=True, color=COLOR_WHITE)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_idx += 1
    
    final_total = pivot.loc['Total general', 'Total general']
    cell = ws.cell(row=total_row, column=total_col_idx, value=int(final_total) if final_total > 0 else "")
    cell.fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
    cell.font = Font(bold=True, color=COLOR_WHITE)
    cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 25
    for i in range(2, total_col_idx + 1):
        ws.column_dimensions[chr(64 + i)].width = 12
    
    print(f"✅ Hoja '{sheet_title}' creada con diseño negro y rojo")


def create_dynamic_vision_diaria_sheet(wb, pivot_main, df):
    """Crea la hoja de Visión Diaria con diseño negro y rojo"""
    
    if 'Visión Diaria' in wb.sheetnames:
        ws = wb['Visión Diaria']
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet('Visión Diaria')
    
    ws.sheet_view.showGridLines = False
    
    border = Border(left=Side(style='thin', color=COLOR_BORDER), 
                   right=Side(style='thin', color=COLOR_BORDER),
                   top=Side(style='thin', color=COLOR_BORDER), 
                   bottom=Side(style='thin', color=COLOR_BORDER))
    
    # Botón volver al menú - Rojo
    ws['A1'] = "⬅ MENÚ"
    ws['A1'].font = Font(bold=True, size=10, color=COLOR_WHITE, underline="single")
    ws['A1'].fill = PatternFill(start_color=COLOR_DARK_RED, end_color=COLOR_DARK_RED, fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].hyperlink = "#'MENÚ GERENCIAL'!A1"
    ws['A1'].border = border
    ws.row_dimensions[1].height = 25
    
    # Fecha de actualización - Negro
    ws['C1'] = "Fecha Actualización"
    ws['C1'].fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
    ws['C1'].font = Font(bold=True, color=COLOR_WHITE)
    ws['C1'].border = border
    ws['C1'].alignment = Alignment(horizontal="center", vertical="center")
    
    ws['D1'] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws['D1'].fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
    ws['D1'].font = Font(color=COLOR_WHITE, bold=True)
    ws['D1'].border = border
    ws['D1'].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.row_dimensions[2].height = 5
    
    # Títulos - Rojo
    ws['A3'] = "Tipo de Trabajo"
    ws['A3'].font = Font(bold=True, size=11, color=COLOR_DARK_RED)
    
    ws['B3'] = "Fecha"
    ws['B3'].font = Font(bold=True, size=11, color=COLOR_DARK_RED)
    
    ws.row_dimensions[4].height = 5
    
    current_row = 5
    
    # Primera fila de encabezados
    ws.cell(row=current_row, column=1, value="")
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
    ws.cell(row=current_row, column=1).border = border
    
    ws.cell(row=current_row, column=2, value="▼ 2025")
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color=COLOR_DARK_RED, end_color=COLOR_DARK_RED, fill_type="solid")
    ws.cell(row=current_row, column=2).font = Font(bold=True, color=COLOR_WHITE)
    ws.cell(row=current_row, column=2).border = border
    ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
    
    months_shown = ['may', 'ago', 'sep', 'oct']
    col_idx = 3
    for month in months_shown:
        ws.cell(row=current_row, column=col_idx, value=f"▼ {month}")
        ws.cell(row=current_row, column=col_idx).fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
        ws.cell(row=current_row, column=col_idx).font = Font(bold=True, color=COLOR_WHITE)
        ws.cell(row=current_row, column=col_idx).border = border
        ws.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
        col_idx += 1
    
    current_row += 1
    
    # Segunda fila de encabezados - Fechas
    ws.cell(row=current_row, column=1, value="")
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
    ws.cell(row=current_row, column=1).border = border
    
    ws.cell(row=current_row, column=2, value="▼")
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
    ws.cell(row=current_row, column=2).border = border
    
    if not pivot_main.empty:
        date_columns = [col for col in pivot_main.columns if col != 'Total general']
        try:
            date_columns_sorted = sorted(date_columns)
        except:
            date_columns_sorted = date_columns
        
        col_idx = 3
        for date_col in date_columns_sorted:
            if date_col != 'Total general':
                try:
                    if hasattr(date_col, 'strftime'):
                        date_str = date_col.strftime("%d/%m/%Y")
                    else:
                        date_str = str(date_col)
                    ws.cell(row=current_row, column=col_idx, value=date_str)
                    ws.cell(row=current_row, column=col_idx).font = Font(color=COLOR_DARK_RED, bold=True)
                    ws.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
                except:
                    ws.cell(row=current_row, column=col_idx, value=str(date_col))
                ws.cell(row=current_row, column=col_idx).border = border
                col_idx += 1
        
        ws.cell(row=current_row, column=col_idx, value="Total general")
        ws.cell(row=current_row, column=col_idx).fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
        ws.cell(row=current_row, column=col_idx).font = Font(bold=True, color=COLOR_WHITE)
        ws.cell(row=current_row, column=col_idx).border = border
        ws.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
        total_col_idx = col_idx
        
        current_row += 1
        
        # Filas de datos - Rojo claro
        for trabajo in pivot_main.index:
            if trabajo != 'Total general':
                ws.cell(row=current_row, column=1, value=trabajo)
                ws.cell(row=current_row, column=1).fill = PatternFill(start_color=COLOR_LIGHT_RED, end_color=COLOR_LIGHT_RED, fill_type="solid")
                ws.cell(row=current_row, column=1).border = border
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                
                ws.cell(row=current_row, column=2, value="")
                ws.cell(row=current_row, column=2).fill = PatternFill(start_color=COLOR_LIGHT_RED, end_color=COLOR_LIGHT_RED, fill_type="solid")
                ws.cell(row=current_row, column=2).border = border
                
                col_idx = 3
                for date_col in date_columns_sorted:
                    value = pivot_main.loc[trabajo, date_col]
                    cell = ws.cell(row=current_row, column=col_idx, value=int(value) if value > 0 else "")
                    cell.fill = PatternFill(start_color=COLOR_LIGHT_RED, end_color=COLOR_LIGHT_RED, fill_type="solid")
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    if value > 0:
                        sheet_name = f"Det_VD_{trabajo[:10]}_{date_col.strftime('%d%m') if hasattr(date_col, 'strftime') else str(date_col)[:10]}"
                        sheet_name = re.sub(r"[:\\/?*\[\]]", "_", sheet_name)[:31]
                        
                        cell.hyperlink = f"#'{sheet_name}'!A1"
                        cell.font = Font(color=COLOR_DARK_RED, underline="single", bold=True)
                        
                        if not hasattr(wb, '_detail_sheets'):
                            wb._detail_sheets = []
                        wb._detail_sheets.append({
                            'sheet_name': sheet_name,
                            'filters': {
                                COL_TIPO_TRABAJO: trabajo,
                                COL_FECHA: date_col
                            }
                        })
                    
                    col_idx += 1
                
                total_value = pivot_main.loc[trabajo, 'Total general']
                cell = ws.cell(row=current_row, column=total_col_idx, value=int(total_value) if total_value > 0 else "")
                cell.fill = PatternFill(start_color=COLOR_LIGHT_RED, end_color=COLOR_LIGHT_RED, fill_type="solid")
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                if total_value > 0:
                    sheet_name = f"Det_VD_{trabajo[:20]}_Total"
                    sheet_name = re.sub(r"[:\\/?*\[\]]", "_", sheet_name)[:31]
                    cell.hyperlink = f"#'{sheet_name}'!A1"
                    cell.font = Font(color=COLOR_DARK_RED, underline="single", bold=True)
                    
                    if not hasattr(wb, '_detail_sheets'):
                        wb._detail_sheets = []
                    wb._detail_sheets.append({
                        'sheet_name': sheet_name,
                        'filters': {COL_TIPO_TRABAJO: trabajo}
                    })
                
                current_row += 1
        
        # Fila de totales - Negro con texto blanco
        total_row = current_row
        ws.cell(row=total_row, column=1, value="Total general")
        ws.cell(row=total_row, column=1).fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
        ws.cell(row=total_row, column=1).font = Font(bold=True, color=COLOR_WHITE)
        ws.cell(row=total_row, column=1).border = border
        
        ws.cell(row=total_row, column=2, value="")
        ws.cell(row=total_row, column=2).fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
        ws.cell(row=total_row, column=2).border = border
        
        col_idx = 3
        for date_col in date_columns_sorted:
            total_value = pivot_main.loc['Total general', date_col]
            cell = ws.cell(row=total_row, column=col_idx, value=int(total_value) if total_value > 0 else "")
            cell.fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
            cell.font = Font(bold=True, color=COLOR_WHITE)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col_idx += 1
        
        final_total = pivot_main.loc['Total general', 'Total general']
        cell = ws.cell(row=total_row, column=total_col_idx, value=int(final_total) if final_total > 0 else "")
        cell.fill = PatternFill(start_color=COLOR_HEADER_BLACK, end_color=COLOR_HEADER_BLACK, fill_type="solid")
        cell.font = Font(bold=True, color=COLOR_WHITE)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    for i in range(3, total_col_idx + 1):
        ws.column_dimensions[chr(64 + i)].width = 11
    
    print(f"✅ Hoja 'Visión Diaria' creada con diseño negro y rojo")


def main():
    """Función principal"""
    output_path = OUTPUT_XLSX
    
    try:
        print("🚀 Iniciando generación de informe diario DINÁMICO CON MENÚ GERENCIAL...")
        print("🎨 Aplicando diseño NEGRO Y ROJO...")
        
        df = fetch_data()
        
        if df.empty:
            print("❌ No se encontraron datos para procesar")
            return False
        
        print("📊 Creando tablas dinámicas...")
        pivot_main, df_processed = build_daily_pivot(df)
        
        # Calcular estadísticas para el menú
        stats = {
            'total_registros': len(df_processed),
            'total_aliados': df_processed[COL_ALIADO].nunique() if COL_ALIADO in df_processed.columns else 0,
            'total_areas': df_processed[COL_AREA].nunique() if COL_AREA in df_processed.columns else 0,
            'total_tipos': df_processed[COL_TIPO_CLIENTE].nunique() if COL_TIPO_CLIENTE in df_processed.columns else 0
        }
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name="Datos Completos", index=False)
        
        print("✅ Archivo base creado")
        
        wb = load_workbook(output_path)
        
        # Crear hojas principales
        print("\n📄 Creando hojas principales...")
        create_dynamic_vision_diaria_sheet(wb, pivot_main, df_processed)
        create_summary_sheet(wb, df_processed, COL_ALIADO, 'Resumen por Aliado', 'Al')
        create_summary_sheet(wb, df_processed, COL_AREA, 'Resumen por Area', 'Ar')
        create_summary_sheet(wb, df_processed, COL_TIPO_CLIENTE, 'Resumen por Tipo Cliente', 'TC')
        
        # Crear hojas de detalle (AHORA VISIBLES)
        if hasattr(wb, '_detail_sheets'):
            print(f"\n📄 Creando {len(wb._detail_sheets)} hojas de detalle (VISIBLES)...")
            created_sheets = set()
            for detail_info in wb._detail_sheets:
                sheet_name = detail_info['sheet_name']
                # Evitar duplicados
                if sheet_name not in created_sheets:
                    try:
                        create_detail_sheet(
                            wb,
                            df_processed,
                            detail_info['filters'].copy(),
                            sheet_name
                        )
                        created_sheets.add(sheet_name)
                    except Exception as e:
                        print(f"⚠️ Error creando hoja {sheet_name}: {e}")
        
        # Crear menú gerencial AL FINAL (para que quede como primera hoja)
        create_menu_sheet(wb, stats)
        
        # Reordenar hojas - Menú primero, luego principales, luego detalles
        sheet_order = ['MENÚ GERENCIAL', 'Visión Diaria', 'Resumen por Aliado', 'Resumen por Area', 'Resumen por Tipo Cliente']
        idx = 0
        for sheet_name in sheet_order:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                wb._sheets.remove(ws)
                wb._sheets.insert(idx, ws)
                idx += 1
        
        wb.save(output_path)
        
        print(f"\n🎉 Proceso completado exitosamente!")
        print(f"📁 Archivo guardado como: {output_path}")
        print(f"📊 Total de hojas generadas: {len(wb.sheetnames)}")
        print(f"\n📋 ESTADÍSTICAS:")
        print(f"   • Total de Cierres: {stats['total_registros']}")
        print(f"   • Total de Aliados: {stats['total_aliados']}")
        print(f"   • Total de Áreas: {stats['total_areas']}")
        print(f"   • Total de Tipos Cliente: {stats['total_tipos']}")
        print(f"\n✨ HOJAS PRINCIPALES:")
        print("   ✓ MENÚ GERENCIAL (Inicio) - 🖤 Negro y ❤️ Rojo")
        print("   ✓ Visión Diaria - 🖤 Negro y ❤️ Rojo")
        print("   ✓ Resumen por Aliado - 🖤 Negro y ❤️ Rojo")
        print("   ✓ Resumen por Area - 🖤 Negro y ❤️ Rojo")
        print("   ✓ Resumen por Tipo Cliente - 🖤 Negro y ❤️ Rojo")
        if hasattr(wb, '_detail_sheets'):
            print(f"\n📄 HOJAS DE DETALLE:")
            print(f"   ✓ {len(created_sheets)} hojas de detalle VISIBLES y funcionales")
        print(f"\n💡 NAVEGACIÓN:")
        print("   • Comience desde el MENÚ GERENCIAL")
        print("   • Haga clic en cualquier número para ver detalles")
        print("   • Las hojas de detalle se abrirán automáticamente ✅")
        print("   • Use el botón '⬅ MENÚ' para volver al inicio")
        print(f"\n🎨 DISEÑO:")
        print("   • Colores estandarizados: Negro y Rojo")
        print("   • Encabezados: Fondo Negro con texto Blanco")
        print("   • Datos: Fondo Rojo claro")
        print("   • Hipervínculos: Rojo oscuro subrayado")
        print("   • Totales: Fondo Negro con texto Blanco")
        
        return True
        
    except Exception as e:
        print(f"❌ Error en el proceso principal: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    main()