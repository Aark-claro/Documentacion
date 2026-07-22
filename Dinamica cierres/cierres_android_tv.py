"""
cierres_android_tv.py
Informe diario de instalaciones Android TV - Región Occidente
Conexión: MySQL remoto vía túnel SSH
BD: otc_backlog / tabla: wf_dia
"""

import re
import warnings
import pandas as pd
import pymysql
from sshtunnel import SSHTunnelForwarder
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

warnings.filterwarnings('ignore')

# ── Configuración SSH ──────────────────────────────────────────────
SSH_HOST     = '186.147.60.119'
SSH_USER     = 'ccot'
SSH_PASSWORD = 'Siesconmigo*'

# ── Configuración MySQL ────────────────────────────────────────────
DB_HOST     = '127.0.0.1'
DB_PORT     = 3307
DB_USER     = 'root'
DB_PASSWORD = '22122012Elf@'
DB_NAME     = 'otc_backlog'

# ── Consulta ───────────────────────────────────────────────────────
SQL_QUERY = """
SELECT *
FROM wf_dia
WHERE `INSTALACION ANDROID TV` = 'si'
  AND Origen = 'REGION OCCIDENTE'
  AND Fecha >= CURDATE()
  AND Fecha <  DATE_ADD(CURDATE(), INTERVAL 1 DAY);
"""

# ── Columnas mapeadas ──────────────────────────────────────────────
COL_TIPO_TRABAJO = "Tipo de Actividad"
COL_FECHA        = "Fecha"
COL_ALIADO       = "Compañia"
COL_AREA         = "Ciudad"
COL_TIPO_RED     = "Tipo de Red"
COL_ESTADO       = "Estado"

OUTPUT_XLSX = "Informe_AndroidTV_Diario.xlsx"

# ── Paleta de colores ──────────────────────────────────────────────
COLOR_BLACK      = "000000"
COLOR_RED_DARK   = "C00000"
COLOR_RED_LIGHT  = "F4B084"
COLOR_WHITE      = "FFFFFF"
COLOR_GRAY_LIGHT = "D9D9D9"
COLOR_GRAY_BORDER= "808080"


# ══════════════════════════════════════════════════════════════════
# CAPA DE DATOS
# ══════════════════════════════════════════════════════════════════

def fetch_data() -> pd.DataFrame:
    """Abre el túnel SSH, ejecuta la consulta y retorna un DataFrame."""
    print("🔐 Abriendo túnel SSH...")
    try:
        with SSHTunnelForwarder(
            (SSH_HOST, 22),
            ssh_username=SSH_USER,
            ssh_password=SSH_PASSWORD,
            remote_bind_address=(DB_HOST, DB_PORT)
        ) as tunnel:
            print(f"✅ Túnel establecido (puerto local: {tunnel.local_bind_port})")
            conn = pymysql.connect(
                host='127.0.0.1',
                port=tunnel.local_bind_port,
                user=DB_USER,
                password=DB_PASSWORD,
                database=DB_NAME,
                charset='utf8mb4'
            )
            df = pd.read_sql(SQL_QUERY, conn)
            conn.close()
            print(f"✅ Datos obtenidos: {len(df)} registros")
            return df
    except Exception as e:
        print(f"❌ Error de conexión: {e}")
        raise


def prepare_data(df: pd.DataFrame) -> pd.DataFrame:
    """Limpia y prepara el DataFrame para los pivotes."""
    df = df.copy()

    # Normalizar fecha
    df[COL_FECHA] = pd.to_datetime(df[COL_FECHA], errors='coerce')
    df = df.dropna(subset=[COL_FECHA])
    df['FECHA_DATE']       = df[COL_FECHA].dt.date
    df['FECHA_FORMATEADA'] = df[COL_FECHA].dt.strftime('%Y-%m-%d')

    # Rellenar nulos en columnas de agrupación
    for col in [COL_TIPO_TRABAJO, COL_ALIADO, COL_AREA, COL_TIPO_RED, COL_ESTADO]:
        if col in df.columns:
            df[col] = df[col].fillna(f'SIN {col}').str.strip()

    return df


def build_pivot(df: pd.DataFrame, index_col: str) -> pd.DataFrame:
    """Crea tabla dinámica: index_col × FECHA_DATE con conteo."""
    pivot = df.pivot_table(
        values='FECHA_FORMATEADA',
        index=index_col,
        columns='FECHA_DATE',
        aggfunc='count',
        fill_value=0
    )
    pivot['Total general'] = pivot.sum(axis=1)
    totals = pivot.sum(axis=0).to_frame().T
    totals.index = ['Total general']
    return pd.concat([pivot, totals])


# ══════════════════════════════════════════════════════════════════
# HELPERS DE ESTILO
# ══════════════════════════════════════════════════════════════════

def _border(style='thin'):
    s = Side(style=style, color=COLOR_GRAY_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

def _style_header(cell, color=COLOR_BLACK, text_color=COLOR_WHITE, size=11):
    cell.fill  = _fill(color)
    cell.font  = Font(bold=True, color=text_color, size=size)
    cell.border = _border()
    cell.alignment = Alignment(horizontal='center', vertical='center')

def _style_data(cell, color=COLOR_RED_LIGHT, bold=False):
    cell.fill   = _fill(color)
    cell.font   = Font(bold=bold)
    cell.border = _border()
    cell.alignment = Alignment(horizontal='center', vertical='center')

def _style_link(cell, color=COLOR_RED_LIGHT):
    cell.fill   = _fill(color)
    cell.font   = Font(color=COLOR_RED_DARK, underline='single', bold=True)
    cell.border = _border()
    cell.alignment = Alignment(horizontal='center', vertical='center')

def _safe_sheet_name(raw: str) -> str:
    """Limpia y trunca un nombre para que sea válido en Excel (máx 31 chars)."""
    return re.sub(r'[:\\/?*\[\]]', '_', raw)[:31]


# ══════════════════════════════════════════════════════════════════
# HOJA: MENÚ GERENCIAL
# ══════════════════════════════════════════════════════════════════

def create_menu_sheet(wb, stats: dict):
    if 'MENÚ GERENCIAL' in wb.sheetnames:
        del wb['MENÚ GERENCIAL']
    ws = wb.create_sheet('MENÚ GERENCIAL', 0)
    ws.sheet_view.showGridLines = False

    for col, w in [('A', 2), ('B', 38), ('C', 38), ('D', 2)]:
        ws.column_dimensions[col].width = w

    # Título principal
    ws.merge_cells('A1:D1')
    ws['A1'] = '📱 DASHBOARD - INSTALACIONES ANDROID TV | REGIÓN OCCIDENTE'
    ws['A1'].font = Font(bold=True, size=16, color=COLOR_WHITE)
    ws['A1'].fill = _fill(COLOR_BLACK)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    # Subtítulo fecha
    ws.merge_cells('A2:D2')
    ws['A2'] = f"Última actualización: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}  |  Filtro: INSTALACION ANDROID TV = 'si'  |  Origen: REGIÓN OCCIDENTE"
    ws['A2'].font = Font(size=10, color=COLOR_RED_DARK, italic=True, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 22

    ws.row_dimensions[3].height = 10

    b = _border('medium')

    # Card total
    ws.merge_cells('B4:C4')
    ws['B4'] = 'RESUMEN GENERAL'
    ws['B4'].font = Font(bold=True, size=13, color=COLOR_WHITE)
    ws['B4'].fill = _fill(COLOR_BLACK)
    ws['B4'].alignment = Alignment(horizontal='center')
    ws['B4'].border = b

    ws.merge_cells('B5:C5')
    ws['B5'] = f"Total Android TV hoy: {stats['total']}"
    ws['B5'].font = Font(size=22, bold=True, color=COLOR_RED_DARK)
    ws['B5'].fill = _fill(COLOR_WHITE)
    ws['B5'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B5'].border = b
    ws.row_dimensions[5].height = 48

    ws.row_dimensions[6].height = 12

    # Accesos rápidos
    ws.merge_cells('B7:C7')
    ws['B7'] = 'ACCESOS RÁPIDOS A REPORTES'
    ws['B7'].font = Font(bold=True, size=13, color=COLOR_WHITE)
    ws['B7'].fill = _fill(COLOR_BLACK)
    ws['B7'].alignment = Alignment(horizontal='center')
    ws['B7'].border = b
    ws.row_dimensions[7].height = 28

    ws.row_dimensions[8].height = 8

    cards = [
        {'title': '📅 VISIÓN DIARIA',        'desc': 'Por Tipo de Actividad × Fecha',  'sheet': 'Visión Diaria'},
        {'title': '🏢 RESUMEN POR ALIADO',   'desc': 'Conteo de OT por empresa aliada','sheet': 'Resumen por Aliado'},
        {'title': '📍 RESUMEN POR CIUDAD',   'desc': 'Distribución geográfica',         'sheet': 'Resumen por Ciudad'},
        {'title': '🌐 RESUMEN POR TIPO RED', 'desc': 'FTTH / Masivo Bidireccional / …', 'sheet': 'Resumen por Tipo Red'},
    ]

    cur = 9
    for idx, card in enumerate(cards):
        col = 'B' if idx % 2 == 0 else 'C'

        ws[f'{col}{cur}'] = card['title']
        ws[f'{col}{cur}'].font = Font(bold=True, size=11, color=COLOR_WHITE)
        ws[f'{col}{cur}'].fill = _fill(COLOR_BLACK)
        ws[f'{col}{cur}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{col}{cur}'].border = b

        ws[f'{col}{cur+1}'] = card['desc']
        ws[f'{col}{cur+1}'].font = Font(size=10, color=COLOR_BLACK)
        ws[f'{col}{cur+1}'].fill = _fill(COLOR_WHITE)
        ws[f'{col}{cur+1}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws[f'{col}{cur+1}'].border = b

        ws[f'{col}{cur+2}'] = '▶ VER REPORTE'
        ws[f'{col}{cur+2}'].font = Font(bold=True, size=10, color=COLOR_WHITE, underline='single')
        ws[f'{col}{cur+2}'].fill = _fill(COLOR_RED_DARK)
        ws[f'{col}{cur+2}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{col}{cur+2}'].border = b
        ws[f'{col}{cur+2}'].hyperlink = f"#'{card['sheet']}'!A1"

        ws.row_dimensions[cur].height   = 28
        ws.row_dimensions[cur+1].height = 32
        ws.row_dimensions[cur+2].height = 26

        if (idx + 1) % 2 == 0:
            cur += 4

    # Info
    cur += 3
    ws.merge_cells(f'B{cur}:C{cur}')
    ws[f'B{cur}'] = 'ℹ️ NAVEGACIÓN'
    ws[f'B{cur}'].font = Font(bold=True, size=11, color=COLOR_WHITE)
    ws[f'B{cur}'].fill = _fill(COLOR_BLACK)
    ws[f'B{cur}'].alignment = Alignment(horizontal='center')
    ws[f'B{cur}'].border = b
    cur += 1

    ws.merge_cells(f'B{cur}:C{cur}')
    ws[f'B{cur}'] = (
        "• Haga clic en cualquier número para ver el detalle\n"
        "• Use el botón '⬅ MENÚ' en cada hoja para volver aquí\n"
        "• Los datos corresponden únicamente a OT con Android TV = 'si'"
    )
    ws[f'B{cur}'].font = Font(size=9, color=COLOR_BLACK)
    ws[f'B{cur}'].fill = _fill(COLOR_GRAY_LIGHT)
    ws[f'B{cur}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws[f'B{cur}'].border = b
    ws.row_dimensions[cur].height = 55

    print("✅ Hoja 'MENÚ GERENCIAL' creada")


# ══════════════════════════════════════════════════════════════════
# HOJA: DETALLE (filtrada, visible)
# ══════════════════════════════════════════════════════════════════

def create_detail_sheet(wb, df_filtered: pd.DataFrame, sheet_name: str):
    """Escribe una hoja de detalle con los datos filtrados."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    b = _border()

    # Encabezado
    ws['A1'] = f'Detalle — {len(df_filtered)} registros'
    _style_header(ws['A1'], size=13)
    ws.row_dimensions[1].height = 28

    # Botón volver
    ws['A2'] = '⬅ MENÚ'
    ws['A2'].font = Font(bold=True, size=10, color=COLOR_WHITE, underline='single')
    ws['A2'].fill = _fill(COLOR_RED_DARK)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'].hyperlink = "#'MENÚ GERENCIAL'!A1"
    ws['A2'].border = b
    ws.row_dimensions[2].height = 22

    HDR_ROW = 4
    # Encabezados de columnas
    for ci, col in enumerate(df_filtered.columns, 1):
        cell = ws.cell(row=HDR_ROW, column=ci, value=col)
        _style_header(cell)

    # Datos
    for ri, row_data in enumerate(df_filtered.itertuples(index=False), HDR_ROW + 1):
        bg = COLOR_RED_LIGHT if ri % 2 == 0 else COLOR_WHITE
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = b
            cell.fill   = _fill(bg)
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # Anchos automáticos
    for col_cells in ws.iter_cols(min_row=HDR_ROW, max_row=HDR_ROW):
        ltr = col_cells[0].column_letter
        max_len = max(
            (len(str(ws.cell(row=r, column=col_cells[0].column).value or ''))
             for r in range(HDR_ROW, ws.max_row + 1)),
            default=8
        )
        ws.column_dimensions[ltr].width = min(max_len + 2, 45)

    # Filtro automático
    last_col = ws.cell(row=HDR_ROW, column=len(df_filtered.columns)).column_letter
    ws.auto_filter.ref = f"A{HDR_ROW}:{last_col}{HDR_ROW + len(df_filtered)}"


# ══════════════════════════════════════════════════════════════════
# HOJA: RESUMEN (pivote genérico)
# ══════════════════════════════════════════════════════════════════

def create_summary_sheet(wb, df: pd.DataFrame, index_col: str,
                         sheet_title: str, prefix: str,
                         detail_registry: list):
    """
    Crea una hoja de resumen pivoteada.
    Registra en detail_registry las hojas de detalle a crear.
    """
    if index_col not in df.columns:
        print(f"⚠️  Columna '{index_col}' no encontrada, saltando '{sheet_title}'")
        return

    if sheet_title in wb.sheetnames:
        del wb[sheet_title]
    ws = wb.create_sheet(sheet_title)
    ws.sheet_view.showGridLines = False

    b = _border()

    # ── Cabecera de hoja ──────────────────────────────────────────
    ws['A1'] = '⬅ MENÚ'
    ws['A1'].font = Font(bold=True, size=10, color=COLOR_WHITE, underline='single')
    ws['A1'].fill = _fill(COLOR_RED_DARK)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].hyperlink = "#'MENÚ GERENCIAL'!A1"
    ws['A1'].border = b
    ws.row_dimensions[1].height = 24

    ws['C1'] = 'Actualización'
    _style_header(ws['C1'])
    ws['D1'] = datetime.now().strftime('%d/%m/%Y %H:%M')
    _style_header(ws['D1'])

    ws['A3'] = sheet_title
    ws['A3'].font = Font(bold=True, size=13, color=COLOR_RED_DARK)
    ws.row_dimensions[4].height = 6

    pivot = build_pivot(df, index_col)
    date_cols = sorted(c for c in pivot.columns if c != 'Total general')

    HDR = 5

    # Encabezado col index
    cell = ws.cell(row=HDR, column=1, value=index_col)
    _style_header(cell)

    # Encabezados fechas
    for ci, dc in enumerate(date_cols, 2):
        cell = ws.cell(row=HDR, column=ci,
                       value=dc.strftime('%d/%m/%Y') if hasattr(dc, 'strftime') else str(dc))
        _style_header(cell)
    total_ci = len(date_cols) + 2
    cell = ws.cell(row=HDR, column=total_ci, value='Total general')
    _style_header(cell)

    # Filas de datos
    cur = HDR + 1
    for item in pivot.index:
        if item == 'Total general':
            continue
        cell = ws.cell(row=cur, column=1, value=item)
        cell.fill   = _fill(COLOR_RED_LIGHT)
        cell.font   = Font(bold=True)
        cell.border = b

        for ci, dc in enumerate(date_cols, 2):
            val = int(pivot.loc[item, dc])
            c = ws.cell(row=cur, column=ci, value=val if val > 0 else '')
            _style_data(c)
            if val > 0:
                sname = _safe_sheet_name(
                    f"Det_{prefix}_{str(item)[:8]}_{dc.strftime('%d%m') if hasattr(dc,'strftime') else str(dc)[:5]}"
                )
                _style_link(c)
                c.value = val
                c.hyperlink = f"#'{sname}'!A1"
                detail_registry.append({
                    'sheet_name': sname,
                    'index_col': index_col,
                    'item': item,
                    'date': dc
                })

        tot = int(pivot.loc[item, 'Total general'])
        c = ws.cell(row=cur, column=total_ci, value=tot if tot > 0 else '')
        _style_data(c)
        if tot > 0:
            sname = _safe_sheet_name(f"Det_{prefix}_{str(item)[:16]}_Tot")
            _style_link(c)
            c.value = tot
            c.hyperlink = f"#'{sname}'!A1"
            detail_registry.append({
                'sheet_name': sname,
                'index_col': index_col,
                'item': item,
                'date': None
            })
        cur += 1

    # Fila de totales
    cell = ws.cell(row=cur, column=1, value='Total general')
    _style_header(cell)
    for ci, dc in enumerate(date_cols, 2):
        val = int(pivot.loc['Total general', dc])
        cell = ws.cell(row=cur, column=ci, value=val if val > 0 else '')
        _style_header(cell)
    gt = int(pivot.loc['Total general', 'Total general'])
    cell = ws.cell(row=cur, column=total_ci, value=gt if gt > 0 else '')
    _style_header(cell)

    # Anchos
    ws.column_dimensions['A'].width = 30
    for i in range(2, total_ci + 1):
        ws.column_dimensions[ws.cell(row=HDR, column=i).column_letter].width = 13

    print(f"✅ Hoja '{sheet_title}' creada")


# ══════════════════════════════════════════════════════════════════
# HOJA: VISIÓN DIARIA (pivote Tipo de Actividad × Fecha)
# ══════════════════════════════════════════════════════════════════

def create_vision_diaria(wb, df: pd.DataFrame, detail_registry: list):
    create_summary_sheet(
        wb, df,
        index_col=COL_TIPO_TRABAJO,
        sheet_title='Visión Diaria',
        prefix='VD',
        detail_registry=detail_registry
    )

# ══════════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL
# ══════════════════════════════════════════════════════════════════

def main():
    print("🚀 Iniciando informe diario Android TV...")

    # 1. Obtener datos
    df_raw = fetch_data()

    if df_raw.empty:
        print("⚠️  No hay datos para hoy. Verifica que la tabla wf_dia esté actualizada.")
        print("    El archivo no será generado.")
        return False

    # 2. Preparar datos
    df = prepare_data(df_raw)
    stats = {'total': len(df)}

    # 3. Crear Excel base (hoja de datos crudos oculta)
    print("📄 Creando archivo Excel base...")
    with pd.ExcelWriter(OUTPUT_XLSX, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='_Datos', index=False)

    wb = load_workbook(OUTPUT_XLSX)
    # Ocultar hoja de datos crudos
    wb['_Datos'].sheet_state = 'hidden'

    # 4. Registro compartido de hojas de detalle
    detail_registry = []

    # 5. Crear hojas de resumen
    print("\n📊 Creando hojas de resumen...")
    create_vision_diaria(wb, df, detail_registry)
    create_summary_sheet(wb, df, COL_ALIADO,    'Resumen por Aliado',   'Al', detail_registry)
    create_summary_sheet(wb, df, COL_AREA,      'Resumen por Ciudad',   'Ci', detail_registry)
    create_summary_sheet(wb, df, COL_TIPO_RED,  'Resumen por Tipo Red', 'TR', detail_registry)

    # 6. Crear hojas de detalle
    seen = set()
    print(f"\n📄 Creando {len(detail_registry)} hojas de detalle...")
    for entry in detail_registry:
        sname = entry['sheet_name']
        if sname in seen:
            continue
        seen.add(sname)

        # Filtrar df según entry
        mask = df[entry['index_col']] == entry['item']
        if entry['date'] is not None:
            mask &= (df['FECHA_DATE'] == entry['date'])
        df_det = df[mask].drop(columns=['FECHA_DATE', 'FECHA_FORMATEADA'], errors='ignore')

        try:
            create_detail_sheet(wb, df_det, sname)
        except Exception as e:
            print(f"  ⚠️  Error en hoja '{sname}': {e}")

    # 7. Crear menú gerencial (al final para que quede primera)
    print("\n🎨 Creando menú gerencial...")
    create_menu_sheet(wb, stats)

    # 8. Ordenar hojas: Menú → resúmenes → datos → detalles
    order = [
        'MENÚ GERENCIAL',
        'Visión Diaria',
        'Resumen por Aliado',
        'Resumen por Ciudad',
        'Resumen por Tipo Red',
        '_Datos'
    ]
    idx = 0
    for name in order:
        if name in wb.sheetnames:
            ws = wb[name]
            wb._sheets.remove(ws)
            wb._sheets.insert(idx, ws)
            idx += 1

    wb.save(OUTPUT_XLSX)

    print(f"\n🎉 Proceso completado.")
    print(f"📁 Archivo: {OUTPUT_XLSX}")
    print(f"📊 Hojas generadas: {len(wb.sheetnames)}")
    print(f"   • Total OT Android TV: {stats['total']}")
    print(f"   • Hojas de detalle:    {len(seen)}")
    return True


if __name__ == '__main__':
    main()
