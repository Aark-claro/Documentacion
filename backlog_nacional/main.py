import csv
import json
import logging
import os
import re
import threading
import time
from datetime import datetime, timedelta
from multiprocessing import cpu_count
from typing import Any, Dict, Optional

import mysql.connector
import polars as pl
import requests
from flask import Flask, jsonify, request
from zoneinfo import ZoneInfo

# Configuración MySQL
DB_CONFIG = {
    'host': 'localhost',
    'port': 3306,
    'user': 'root',
    'password': '123456',
    'database': 'backlog_nacional',
    'charset': 'utf8mb4',
    'autocommit': True,
    'connection_timeout': 60  # Timeout de conexión inicial aumentado
}

# Configuración de Red - Accesible desde toda la red
FLASK_HOST = '0.0.0.0'  # Permite acceso desde cualquier IP de la red
FLASK_PORT = 1211

# Configurar logging
import sys
_file_handler = logging.FileHandler('sharepoint_monitor.log', encoding='utf-8')
_stream_handler = logging.StreamHandler(sys.stdout)
_stream_handler.stream = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1, closefd=False)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        _file_handler,
        _stream_handler,
    ]
)

# Zona horaria de Colombia
COLOMBIA_TZ = ZoneInfo("America/Bogota")

# Configuración SharePoint
SHAREPOINT_BASE = "https://claromovilco.sharepoint.com"
SHAREPOINT_ROOT = "/sites/AseguramientoOperacionCGO/BACKLOG REPORTE DIARIO"
DOWNLOAD_DIR = os.path.join(".", "descargas")
STATE_FILE = "last_download.json"
COOKIES_FILE = "cookies.json"
SHAREPOINT_ALLOWED_PREFIXES = ("Backlog_Nacional",)
SHAREPOINT_MAX_RETRIES = 5
SHAREPOINT_REQUEST_TIMEOUT = 20

# Configuración Backlog Nacional (CSV)
BACKLOG_DB_NAME = "backlog_nacional"
BACKLOG_TABLE_NAME = "backlog_nacional"
BACKLOG_DB_BATCH_SIZE = 5000
BACKLOG_DB_CONFIG = {
    "host": os.environ.get("BACKLOG_DB_HOST", "localhost"),
    "port": int(os.environ.get("BACKLOG_DB_PORT", 3306)),
    "user": os.environ.get("BACKLOG_DB_USER", "root"),
    "password": os.environ.get("BACKLOG_DB_PASSWORD", "123456"),
    "database": BACKLOG_DB_NAME,
    "raise_on_warnings": False,
}

CSV_COLUMN_MAPPING = [
    ("TIPO_TRABAJO", "TIPO_TRABAJO"),
    ("TIPO_BACKLOG", "TIPO_BACKLOG"),
    ("CUENTA", "CUENTA"),
    ("OT/LL", "OT_LL"),
    ("FECHA_CREADO", "FECHA_CREADO"),
    ("HORA_CREADO", "HORA_CREADO"),
    ("USUARIO_CREADOR", "USUARIO_CREADOR"),
    ("NODO", "NODO"),
    ("RAZON", "RAZON"),
    ("SEGMENTO", "SEGMENTO"),
    ("CLASE", "CLASE"),
    ("CONVENIENCIA", "CONVENIENCIA"),
    ("SUS_NOMBRE", "SUS_NOMBRE"),
    ("SUS_APELLIDO", "SUS_APELLIDO"),
    ("TELEFONO1", "TELEFONO1"),
    ("TELEFONO2", "TELEFONO2"),
    ("TELEFONO3", "TELEFONO3"),
    ("DIRECCION_CARRERA", "DIRECCION_CARRERA"),
    ("DIRECCION_CALLE", "DIRECCION_CALLE"),
    ("DIRECCION_APT", "DIRECCION_APT"),
    ("CEDULA_VENDEDOR", "CEDULA_VENDEDOR"),
    ("TARIFA", "TARIFA"),
    ("DIRECCION_CALLE_OLD", "DIRECCION_CALLE_OLD"),
    ("DIRECCION_APT_OLD", "DIRECCION_APT_OLD"),
    ("ANTIGUEDAD_DIGITACION", "ANTIGUEDAD_DIGITACION"),
    ("Region", "REGION"),
    ("Area", "AREA"),
    ("Red", "RED"),
    ("Comunidad", "COMUNIDAD"),
    ("Aliado Zonificado", "ALIADO_ZONIFICADO"),
    ("Opera", "OPERA"),
    ("FECHA_AGENDA_FUTURO", "FECHA_AGENDA_FUTURO"),
    ("FECHA_ULTIMA_VISITA", "FECHA_ULTIMA_VISITA"),
    ("ESTADO_VISITA", "ESTADO_VISITA"),
    ("RAZON_NODONE", "RAZON_NODONE"),
    ("ANTIGUEDAD_ULTIMA_VISITA", "ANTIGUEDAD_ULTIMA_VISITA"),
    ("DIVISION", "DIVISION"),
    ("REGIONAL", "REGIONAL"),
    ("CODIGO_DX", "CODIGO_DX"),
    ("ESTADO_ORDEN", "ESTADO_ORDEN"),
    ("WMTYPS", "WMTYPS"),
    ("CUENTA_MATRIZ", "CUENTA_MATRIZ"),
    ("TRONCAL", "TRONCAL"),
    ("TRONCAL_P", "TRONCAL_P"),
    ("COORDENADAS", "COORDENADAS"),
    ("NOTA_RR_1", "NOTA_RR_1"),
    ("NOTA_RR_2", "NOTA_RR_2"),
]

BACKLOG_DB_COLUMN_ORDER = [dest for _, dest in CSV_COLUMN_MAPPING]
BACKLOG_DB_INSERT_SQL = (
    f"INSERT IGNORE INTO {BACKLOG_TABLE_NAME} ({', '.join(BACKLOG_DB_COLUMN_ORDER)}) "
    f"VALUES ({', '.join(['%s'] * len(BACKLOG_DB_COLUMN_ORDER))})"
)

BACKLOG_DATE_COLUMNS = {"FECHA_CREADO"}
BACKLOG_DATETIME_COLUMNS = {"FECHA_AGENDA_FUTURO", "FECHA_ULTIMA_VISITA"}
BACKLOG_INT_COLUMNS = {"ANTIGUEDAD_DIGITACION", "ANTIGUEDAD_ULTIMA_VISITA"}

def now_colombia() -> datetime:
    """Obtener la hora actual en la zona horaria de Colombia."""
    return datetime.now(COLOMBIA_TZ)

# Variables globales
last_processed_filename = None
consecutive_errors = 0
max_consecutive_errors = 5
last_update_time = None  # Hora de la última actualización exitosa
script_start_time = now_colombia()  # Hora de inicio del script
last_file_date = None  # Fecha del último archivo procesado
total_files_processed = 0  # Contador total de archivos procesados
last_error_message = None  # Último mensaje de error


# Utilidades SharePoint
def ensure_download_folder():
    """Crear carpeta de descargas si no existe."""
    try:
        os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    except OSError as exc:
        logging.error(f"❌ No se pudo crear el directorio de descargas {DOWNLOAD_DIR}: {exc}")
        raise


def load_state() -> Dict[str, str]:
    """Leer estado local con el último archivo descargado/procesado."""
    default_state = {"last_file": "", "last_processed": ""}
    if not os.path.exists(STATE_FILE):
        return default_state.copy()

    try:
        with open(STATE_FILE, "r", encoding="utf-8") as state_file:
            data = json.load(state_file)
        if not isinstance(data, dict):
            return default_state.copy()
        data.setdefault("last_file", "")
        data.setdefault("last_processed", "")
        return data
    except Exception as exc:
        logging.warning(f"⚠️ No se pudo leer {STATE_FILE}: {exc}")
        return default_state.copy()


def save_state(last_file: str, processed_filename: Optional[str] = None):
    """Persistir estado local."""
    state = load_state()
    state["last_file"] = last_file
    if processed_filename is not None:
        state["last_processed"] = processed_filename

    try:
        with open(STATE_FILE, "w", encoding="utf-8") as state_file:
            json.dump(state, state_file)
    except Exception as exc:
        logging.warning(f"⚠️ No se pudo guardar {STATE_FILE}: {exc}")


def chunked(iterable, size):
    """Agrupar elementos en lotes."""
    buffer = []
    for item in iterable:
        buffer.append(item)
        if len(buffer) >= size:
            yield buffer
            buffer = []
    if buffer:
        yield buffer


def try_parse_datetime(value: str, formats) -> Optional[datetime]:
    for fmt in formats:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def normalize_backlog_value(csv_col: str, raw_value: Optional[str]) -> Optional[Any]:
    if raw_value is None:
        return None

    normalized = raw_value.strip()
    if normalized == "":
        return None

    if csv_col in BACKLOG_DATE_COLUMNS:
        parsed = try_parse_datetime(normalized, ("%Y-%m-%d", "%d/%m/%Y"))
        return parsed.date() if parsed else None

    if csv_col in BACKLOG_DATETIME_COLUMNS:
        return try_parse_datetime(normalized, ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"))

    if csv_col in BACKLOG_INT_COLUMNS:
        try:
            return int(normalized)
        except ValueError:
            return None

    return normalized


def read_backlog_rows(file_path: str):
    with open(file_path, newline="", encoding="latin-1") as csvfile:
        reader = csv.DictReader(csvfile)
        for raw in reader:
            row = []
            for csv_col, _ in CSV_COLUMN_MAPPING:
                row.append(normalize_backlog_value(csv_col, raw.get(csv_col)))
            yield tuple(row)


def open_backlog_db_connection():
    try:
        return mysql.connector.connect(**BACKLOG_DB_CONFIG)
    except mysql.connector.Error as exc:
        logging.error(f"❌ No se pudo conectar a MySQL Backlog → {exc}")
        return None


def extraer_fecha_de_archivo(nombre_archivo: str) -> Optional[datetime]:
    """Extraer fecha/hora del nombre del archivo CSV.
    Formato esperado: Backlog_Nacional_Por_Produccion_DD-MM-YYYY_HH.MM.AM/PM.csv
    """
    try:
        # Buscar patrón: DD-MM-YYYY_HH.MM.AM/PM
        match = re.search(r'(\d{2})-(\d{2})-(\d{4})_(\d{2})\.(\d{2})\.(AM|PM)', nombre_archivo)
        if not match:
            return None
        
        dia, mes, anio, hora, minuto, ampm = match.groups()
        hora_int = int(hora)
        
        # Convertir a formato 24 horas
        if ampm == 'PM' and hora_int != 12:
            hora_int += 12
        elif ampm == 'AM' and hora_int == 12:
            hora_int = 0
        
        return datetime(int(anio), int(mes), int(dia), hora_int, int(minuto), 0)
    except Exception:
        return None


def guardar_ultima_actualizacion(nombre_archivo: str):
    """Guardar fecha/hora y nombre del último archivo procesado en la DB."""
    connection = open_backlog_db_connection()
    if not connection:
        logging.warning("⚠️ No se pudo guardar última actualización en DB")
        return
    
    # Extraer fecha del nombre del archivo
    fecha_archivo = extraer_fecha_de_archivo(nombre_archivo)
    if not fecha_archivo:
        logging.warning(f"⚠️ No se pudo extraer fecha del archivo: {nombre_archivo}, usando NOW()")
        fecha_sql = "NOW()"
        params = (nombre_archivo, nombre_archivo)
    else:
        fecha_sql = "%s"
        params = (fecha_archivo, nombre_archivo, fecha_archivo, nombre_archivo)
    
    try:
        cursor = connection.cursor()
        # Crear tabla si no existe
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS ultima_actualizacion (
                id INT PRIMARY KEY DEFAULT 1,
                fecha_descarga DATETIME,
                nombre_archivo VARCHAR(255)
            )
        """)
        # Insertar o actualizar el registro
        if fecha_archivo:
            cursor.execute("""
                INSERT INTO ultima_actualizacion (id, fecha_descarga, nombre_archivo)
                VALUES (1, %s, %s)
                ON DUPLICATE KEY UPDATE fecha_descarga = %s, nombre_archivo = %s
            """, params)
        else:
            cursor.execute("""
                INSERT INTO ultima_actualizacion (id, fecha_descarga, nombre_archivo)
                VALUES (1, NOW(), %s)
                ON DUPLICATE KEY UPDATE fecha_descarga = NOW(), nombre_archivo = %s
            """, (nombre_archivo, nombre_archivo))
        connection.commit()
        logging.info(f"📝 Última actualización guardada en DB: {nombre_archivo} ({fecha_archivo})")
    except Exception as exc:
        logging.warning(f"⚠️ Error guardando última actualización: {exc}")
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()


def process_backlog_file(file_path: str) -> bool:
    """Procesar archivo CSV Backlog Nacional y subirlo a MySQL."""
    logging.info(f"\n🗂 Procesando Backlog Nacional: {file_path}")
    connection = open_backlog_db_connection()
    if not connection:
        return False

    cursor = None
    rows_inserted = 0
    rows_attempted = 0
    try:
        cursor = connection.cursor()
        logging.info("🗑️ Limpiando la tabla `backlog_nacional` antes de volver a poblarla")
        cursor.execute("DELETE FROM backlog_nacional")
        connection.commit()
        for batch in chunked(read_backlog_rows(file_path), BACKLOG_DB_BATCH_SIZE):
            cursor.executemany(BACKLOG_DB_INSERT_SQL, batch)
            inserted = cursor.rowcount
            rows_inserted += inserted
            rows_attempted += len(batch)
            skipped = len(batch) - inserted
            logging.info(f"   ▸ Lote de {len(batch)} → insertados {inserted}, omitidos {skipped}")
        connection.commit()
        logging.info(
            f"✔ Se subieron {rows_inserted} filas a `{BACKLOG_TABLE_NAME}` "
            f"(omitidas {rows_attempted - rows_inserted})."
        )
        # Guardar última actualización en DB
        nombre_archivo = os.path.basename(file_path)
        guardar_ultima_actualizacion(nombre_archivo)
        
        # Enviar reporte de REGION OCCIDENTE por email automáticamente
        try:
            logging.info("📧 Enviando reporte de REGION OCCIDENTE por email...")
            send_region_occidente_report()
        except Exception as email_exc:
            logging.warning(f"⚠️ Error enviando email de REGION OCCIDENTE: {email_exc}")
        
        return True
    except Exception as exc:
        if connection:
            connection.rollback()
        logging.error(f"❌ Error subiendo datos a MySQL Backlog → {exc}")
        return False
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except Exception as exc_rm:
                logging.warning(f"⚠️ No se pudo eliminar archivo CSV temporal: {exc_rm}")


def safe_request(session: requests.Session, url: str) -> Optional[requests.Response]:
    """Petición GET con reintentos para SharePoint."""
    for attempt in range(1, SHAREPOINT_MAX_RETRIES + 1):
        try:
            response = session.get(url, timeout=SHAREPOINT_REQUEST_TIMEOUT)
            if response.status_code == 401:
                logging.error("❌ Sesión inválida (401) al consultar SharePoint. Verifica cookies.json")
                return None
            response.raise_for_status()
            return response
        except Exception as exc:
            wait_time = min(attempt * 5, 60)
            logging.warning(f"⚠️ Error consultando SharePoint: {exc} (intento {attempt}/{SHAREPOINT_MAX_RETRIES})")
            logging.info(f"⏳ Reintentando en {wait_time} segundos...")
            time.sleep(wait_time)
    logging.error("❌ Error persistente consultando SharePoint")
    return None


def create_sharepoint_session() -> Optional[requests.Session]:
    """Crear sesión autenticada con SharePoint usando cookies FedAuth y rtFa."""
    try:
        with open(COOKIES_FILE, "r", encoding="utf-8") as cookies_file:
            cookies = json.load(cookies_file)
    except Exception as exc:
        logging.error(f"❌ No se pudo leer {COOKIES_FILE}: {exc}")
        return None

    fed_auth = cookies.get("FedAuth")
    rt_fa = cookies.get("rtFa")
    if not fed_auth or not rt_fa:
        logging.error("❌ cookies.json no contiene FedAuth/rtFa válidos")
        return None

    session = requests.Session()
    session.headers.update({
        "Accept": "application/json;odata=verbose",
        "User-Agent": "Mozilla/5.0"
    })
    session.cookies.set("FedAuth", fed_auth, domain="claromovilco.sharepoint.com")
    session.cookies.set("rtFa", rt_fa, domain="claromovilco.sharepoint.com")
    return session


def _is_allowed_filename(filename: str) -> bool:
    lowered = filename.lower()
    return any(lowered.startswith(prefix.lower()) for prefix in SHAREPOINT_ALLOWED_PREFIXES)


def get_latest_sharepoint_file(session: requests.Session) -> Optional[Dict[str, Any]]:
    """Obtener el archivo más reciente disponible en SharePoint."""
    logging.info("📌 Buscando archivo más reciente en SharePoint...")

    years_url = (
        f"{SHAREPOINT_BASE}/sites/AseguramientoOperacionCGO/_api/web"
        f"/GetFolderByServerRelativeUrl('{SHAREPOINT_ROOT}')/Folders?$select=Name,ServerRelativeUrl"
    )
    resp = safe_request(session, years_url)
    if not resp:
        return None

    try:
        years_data = resp.json()
        year_folders = [folder for folder in years_data["d"]["results"] if re.fullmatch(r"\d{4}", folder["Name"])]
    except Exception as exc:
        logging.error(f"❌ Error interpretando los años en SharePoint: {exc}")
        return None

    if not year_folders:
        logging.warning("⚠️ No se encontraron carpetas de año en SharePoint")
        return None

    # Ordenar años de más reciente a más antiguo
    sorted_years = sorted(year_folders, key=lambda x: x["Name"], reverse=True)
    
    # Intentar cada año hasta encontrar uno con datos
    for year_folder_info in sorted_years:
        year_folder = year_folder_info["Name"]
        logging.info(f"📁 Verificando año: {year_folder}")

        months_url = (
            f"{SHAREPOINT_BASE}/sites/AseguramientoOperacionCGO/_api/web"
            f"/GetFolderByServerRelativeUrl('{SHAREPOINT_ROOT}/{year_folder}')/Folders?$select=Name,ServerRelativeUrl"
        )
        resp = safe_request(session, months_url)
        if not resp:
            continue

        try:
            months_data = resp.json()
            month_folders = [folder for folder in months_data["d"]["results"] if re.match(r"\d{2}\.", folder["Name"])]
        except Exception as exc:
            logging.error(f"❌ Error interpretando los meses en SharePoint: {exc}")
            continue

        if not month_folders:
            logging.info(f"📁 Año {year_folder} sin carpetas de mes, probando año anterior...")
            continue

        # Ordenar meses de más reciente a más antiguo
        sorted_months = sorted(month_folders, key=lambda x: int(x["Name"].split(".")[0]), reverse=True)
        
        # Intentar cada mes hasta encontrar uno con archivos
        for month_folder_info in sorted_months:
            month_folder = month_folder_info["Name"]
            
            final_folder = f"{SHAREPOINT_ROOT}/{year_folder}/{month_folder}"
            files_url = (
                f"{SHAREPOINT_BASE}/sites/AseguramientoOperacionCGO/_api/web"
                f"/GetFolderByServerRelativeUrl('{final_folder}')"
                "/Files?$select=Name,ServerRelativeUrl,TimeLastModified"
            )
            resp = safe_request(session, files_url)
            if not resp:
                continue

            try:
                files_data = resp.json()
                files = files_data["d"]["results"]
            except Exception as exc:
                logging.error(f"❌ Error interpretando archivos en SharePoint: {exc}")
                continue

            filtered_files = [file_info for file_info in files if _is_allowed_filename(file_info.get("Name", ""))]
            if not filtered_files:
                continue

            # Encontramos archivos válidos
            logging.info(f"📁 Año seleccionado: {year_folder}")
            logging.info(f"📁 Mes seleccionado: {month_folder}")
            
            latest_file = sorted(
                filtered_files,
                key=lambda x: x.get("TimeLastModified") or "",
                reverse=True
            )[0]
            logging.info(f"📄 Archivo encontrado: {latest_file['Name']}")
            return latest_file

    logging.warning("⚠️ No se encontraron archivos compatibles en ningún año/mes de SharePoint")
    return None


def download_sharepoint_file(session: requests.Session, file_info: Dict[str, Any]) -> Optional[str]:
    """Descargar archivo de SharePoint y retornarlo en disco."""
    filename = file_info.get("Name")
    server_relative_url = file_info.get("ServerRelativeUrl")
    if not filename or not server_relative_url:
        logging.error("❌ Información de archivo inválida")
        return None

    ensure_download_folder()
    destination = os.path.join(DOWNLOAD_DIR, filename)
    file_url = f"{SHAREPOINT_BASE}{server_relative_url}"
    logging.info(f"⬇ Descargando desde SharePoint: {filename}")

    for attempt in range(1, SHAREPOINT_MAX_RETRIES + 1):
        try:
            with session.get(file_url, stream=True, timeout=SHAREPOINT_REQUEST_TIMEOUT) as response:
                response.raise_for_status()
                with open(destination, "wb") as out_file:
                    for chunk in response.iter_content(chunk_size=8192):
                        if chunk:
                            out_file.write(chunk)
            logging.info(f"✅ Archivo guardado en {destination}")
            return destination
        except Exception as exc:
            wait_time = min(attempt * 5, 60)
            logging.warning(f"⚠️ Error descargando {filename}: {exc} (intento {attempt}/{SHAREPOINT_MAX_RETRIES})")
            logging.info(f"⏳ Reintentando en {wait_time} segundos...")
            time.sleep(wait_time)

    logging.error(f"❌ No se pudo descargar {filename} después de varios intentos")
    return None


def parse_sharepoint_timestamp(timestamp: Optional[str]) -> Optional[datetime]:
    """Convertir TimeLastModified de SharePoint a datetime con zona horaria de Colombia."""
    if not timestamp:
        return None
    try:
        cleaned = timestamp.replace("Z", "+00:00")
        dt = datetime.fromisoformat(cleaned)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        return dt.astimezone(COLOMBIA_TZ)
    except Exception:
        return None

# Configuración Flask
app = Flask(__name__)
flask_port = 1211

def check_bot_health() -> bool:
    """
    Verificar la salud del bot
    Retorna True si el bot está saludable, False si necesita atención
    """
    global last_update_time
    
    current_time = now_colombia()
    
    # Si nunca se ha actualizado, verificar cuánto tiempo lleva el script corriendo
    if not last_update_time:
        time_since_start = current_time - script_start_time
        minutes_since_start = int(time_since_start.total_seconds() / 60)
        if minutes_since_start > 40:
            logging.warning(f"⚠️ Bot sin ninguna actualización desde hace {minutes_since_start} minutos")
            return False
        return True
    
    # Calcular tiempo desde última actualización
    time_since_update = current_time - last_update_time
    minutes_since_update = int(time_since_update.total_seconds() / 60)
    
    # Umbral de 40 minutos sin actualizaciones
    ALERT_THRESHOLD_MINUTES = 40
    
    # Si ha pasado más tiempo del umbral, registrar en log
    if minutes_since_update > ALERT_THRESHOLD_MINUTES:
        logging.warning(f"⚠️ Bot sin actualizar por {minutes_since_update} minutos")
        logging.warning(f"📊 Última actualización: {last_update_time.strftime('%Y-%m-%d %H:%M:%S')}")
        logging.warning(f"📄 No se han descargado archivos SharePoint con prefijo back_ en {minutes_since_update} minutos")
        return False
    
    return True


def download_and_process_file(part, filename: str) -> bool:
    """
    Descargar archivo desde un correo (payload MIME) y procesarlo.
    """
    if not filename or not filename.lower().endswith((".xlsx", ".xls")):
        logging.error(f"❌ Archivo no válido: {filename}")
        return False

    ensure_download_folder()
    download_path = os.path.join(DOWNLOAD_DIR, filename)
    max_retries = 3

    try:
        logging.info(f"📥 Descargando archivo: {filename}")
        for attempt in range(max_retries):
            try:
                payload = part.get_payload(decode=True)
                if not payload:
                    raise Exception("Payload del archivo está vacío")

                if len(payload) < 100:
                    raise Exception(f"Archivo demasiado pequeño: {len(payload)} bytes")

                with open(download_path, "wb") as downloaded:
                    downloaded.write(payload)

                if not os.path.exists(download_path):
                    raise Exception("El archivo no se guardó correctamente")

                file_size = os.path.getsize(download_path)
                if file_size == 0:
                    raise Exception("El archivo guardado está vacío")

                logging.info(f"📥 Archivo descargado exitosamente: {filename} ({file_size} bytes)")
                break
            except Exception as exc:
                logging.warning(f"⚠️ Error en descarga (intento {attempt + 1}/{max_retries}): {exc}")
                if attempt == max_retries - 1:
                    logging.error("❌ Falló la descarga después de todos los intentos")
                    if os.path.exists(download_path):
                        os.remove(download_path)
                    return False
                time.sleep(2)

        return process_downloaded_file(download_path, filename)
    except Exception as exc:
        logging.error(f"🚨 ERROR CRÍTICO en descarga: {type(exc).__name__}: {exc}")
        if os.path.exists(download_path):
            try:
                os.remove(download_path)
            except Exception:
                pass
        return False


def process_downloaded_file(download_path: str, filename: str) -> bool:
    """
    Procesar un archivo ya descargado ejecutando importación a MySQL.
    """
    conn = None
    max_retries = 3

    try:
        if not os.path.exists(download_path):
            logging.error(f"❌ El archivo {download_path} no existe para procesamiento")
            return False

        file_size = os.path.getsize(download_path)
        if file_size < 100:
            logging.error(f"❌ El archivo {filename} es demasiado pequeño ({file_size} bytes)")
            return False

        df = None
        for attempt in range(max_retries):
            try:
                df = read_excel_file(download_path)
                if df is not None and not df.is_empty():
                    logging.info(f"📊 Excel leído exitosamente: {df.height} filas, {df.width} columnas")
                    break
                raise Exception("DataFrame está vacío o es None")
            except Exception as exc:
                logging.warning(f"⚠️ Error leyendo Excel (intento {attempt + 1}/{max_retries}): {exc}")
                if attempt == max_retries - 1:
                    logging.error("❌ No se pudo leer el archivo Excel después de todos los intentos")
                    return False
                time.sleep(1)

        if df is None or df.is_empty():
            logging.error("❌ No se pudieron obtener datos del archivo Excel")
            return False

        for attempt in range(max_retries):
            try:
                conn = connect_to_mysql()
                if conn:
                    logging.info("🔗 Conexión a MySQL establecida")
                    break

                logging.info("🔄 Intentando conexión simple como fallback...")
                conn = connect_to_mysql_simple()
                if conn:
                    logging.info("🔗 Conexión MySQL simple establecida")
                    break

                raise Exception("No se pudo establecer conexión a MySQL")
            except Exception as exc:
                logging.warning(f"⚠️ Error conectando a MySQL (intento {attempt + 1}/{max_retries}): {exc}")
                if attempt == max_retries - 1:
                    logging.error("❌ No se pudo conectar a MySQL después de todos los intentos")
                    return False
                time.sleep(3)

        success = False
        for attempt in range(max_retries):
            try:
                success = insert_data_to_mysql(conn, df)
                if success:
                    logging.info(f"✅ Datos procesados correctamente: {df.height} registros insertados")
                    break
                raise Exception("Falló la inserción de datos")
            except Exception as exc:
                logging.warning(f"⚠️ Error insertando datos (intento {attempt + 1}/{max_retries}): {exc}")
                if attempt == max_retries - 1:
                    logging.error("❌ No se pudieron insertar los datos después de todos los intentos")
                    success = False
                time.sleep(2)

        return success
    except Exception as exc:
        logging.error(f"🚨 ERROR CRÍTICO en procesamiento: {type(exc).__name__}: {exc}")
        return False
    finally:
        if conn:
            try:
                conn.close()
                logging.info("🔌 Conexión MySQL cerrada")
            except Exception as exc_close:
                logging.warning(f"⚠️ Error cerrando conexión MySQL: {exc_close}")

        if os.path.exists(download_path):
            try:
                os.remove(download_path)
                logging.info("🗑️ Archivo temporal eliminado")
            except Exception as exc_rm:
                logging.warning(f"⚠️ Error eliminando archivo temporal: {exc_rm}")

def read_excel_file(filepath: str) -> Optional[pl.DataFrame]:
    """Leer archivo Excel optimizado"""
    try:
        df = pl.read_excel(filepath, sheet_name='back_data')
        
        # Filtrar y limpiar columnas
        df = df.select([col for col in df.columns if col is not None and str(col).strip() != ''])
        
        # Limpiar nombres de columnas
        new_columns = [str(col).lower().replace(' ', '_').replace('.', '') for col in df.columns]
        df = df.rename(dict(zip(df.columns, new_columns)))
        
        # Filtrar columnas inválidas
        df = df.select([col for col in df.columns if col not in ['nan', '']])
        
        # Manejo de nulls
        if 'tipo_cliente' in df.columns:
            df = df.with_columns(pl.col('tipo_cliente').fill_null(''))
        
        logging.info(f"📊 Excel procesado: {df.height} filas, {df.width} columnas")
        return df
        
    except Exception as e:
        logging.error(f"❌ Error leyendo Excel: {e}")
        return None

def connect_to_mysql(max_retries: int = 5):
    """
    Conectar a MySQL con reintentos automáticos y reconexión robusta
    Sistema de alta disponibilidad que NUNCA falla
    """
    import mysql.connector
    from mysql.connector import Error
    
    base_wait_time = 5  # Reducido de 10 a 5 segundos
    
    for attempt in range(max_retries):
        conn = None
        try:
            logging.info(f"🔗 Conectando a MySQL (intento {attempt + 1}/{max_retries})")
            
            # Configuración simplificada de conexión
            config = DB_CONFIG.copy()
            config.update({
                'autocommit': False,
                'use_unicode': True,
                'raise_on_warnings': False,
                'buffered': True
            })
            
            # Crear conexión
            conn = mysql.connector.connect(**config)
            
            # Verificar que la conexión esté realmente funcionando
            if conn.is_connected():
                cursor = conn.cursor()
                cursor.execute("SELECT 1")
                result = cursor.fetchone()
                cursor.close()
                
                if result and result[0] == 1:
                    logging.info("✅ Conexión a MySQL exitosa y verificada")
                    return conn
                else:
                    raise Exception("La verificación de conexión falló")
            else:
                raise Exception("La conexión no está activa")
                
        except mysql.connector.Error as e:
            error_code = e.errno if hasattr(e, 'errno') else None
            error_msg = str(e).upper()
            
            if conn:
                try:
                    conn.close()
                except:
                    pass
            
            # Errores específicos de MySQL
            if error_code == 1045:  # Access denied
                logging.error("🚨 ERROR CRÍTICO: Credenciales de MySQL inválidas")
                logging.error("🔧 Verifica DB_USER y DB_PASSWORD en las variables de entorno")
                return None
                
            elif error_code == 2003:  # Can't connect to MySQL server
                logging.warning(f"🌐 No se puede conectar al servidor MySQL en intento {attempt + 1}")
                
            elif error_code == 1049:  # Unknown database
                logging.error("🚨 ERROR CRÍTICO: Base de datos no existe")
                logging.error(f"🔧 Verifica que la base de datos '{DB_CONFIG.get('database', 'N/A')}' exista")
                return None
                
            elif error_code == 2006:  # MySQL server has gone away
                logging.warning(f"🔄 Servidor MySQL desconectado en intento {attempt + 1}")
                
            elif 'TIMEOUT' in error_msg or 'TIMED OUT' in error_msg:
                logging.warning(f"⏰ Timeout de conexión MySQL en intento {attempt + 1}")
                
            else:
                logging.warning(f"❌ Error MySQL ({error_code}): {e}")
            
            # Espera entre reintentos para errores de MySQL
            if attempt < max_retries - 1:
                if 'TIMEOUT' in error_msg or 'TIMED OUT' in error_msg:
                    wait_time = 2 * (attempt + 1)  # 2, 4, 6, 8 segundos para timeouts
                else:
                    wait_time = base_wait_time * (2 ** attempt)  # 5, 10, 20, 40 segundos
                wait_time = min(wait_time, 60)  # Máximo 1 minuto
                logging.info(f"⏳ Esperando {wait_time} segundos antes del siguiente intento...")
                time.sleep(wait_time)
                
        except Exception as e:
            if conn:
                try:
                    conn.close()
                except:
                    pass
                    
            logging.warning(f"❌ Error inesperado conectando a MySQL: {type(e).__name__}: {e}")
            
            # Espera más corta entre reintentos para timeouts
            if attempt < max_retries - 1:
                # Para timeouts, esperar menos tiempo
                if 'TIMEOUT' in str(e).upper() or 'TIMED OUT' in str(e).upper():
                    wait_time = 2 * (attempt + 1)  # 2, 4, 6, 8 segundos
                else:
                    wait_time = base_wait_time * (2 ** attempt)  # 5, 10, 20, 40 segundos
                wait_time = min(wait_time, 60)  # Máximo 1 minuto
                logging.info(f"⏳ Esperando {wait_time} segundos antes del siguiente intento...")
                time.sleep(wait_time)
    
    logging.error("🚨 FALLO CRÍTICO: No se pudo conectar a MySQL después de todos los intentos")
    return None

def verify_connection(conn) -> bool:
    """Verificar que la conexión MySQL esté activa"""
    try:
        if not conn or not conn.is_connected():
            return False
        
        cursor = conn.cursor()
        cursor.execute("SELECT 1")
        result = cursor.fetchone()
        cursor.close()
        return result and result[0] == 1
    except:
        return False

def connect_to_mysql_simple():
    """Función de conexión MySQL simplificada como fallback"""
    try:
        # Configuración mínima para conexión rápida
        simple_config = {
            'host': DB_CONFIG['host'],
            'user': DB_CONFIG['user'],
            'password': DB_CONFIG['password'],
            'database': DB_CONFIG['database'],
            'connection_timeout': 10,  # Timeout muy corto
            'autocommit': True
        }
        
        conn = mysql.connector.connect(**simple_config)
        if conn and conn.is_connected():
            logging.info("✅ Conexión MySQL simple exitosa")
            return conn
    except Exception as e:
        logging.warning(f"❌ Conexión simple falló: {e}")
    
    return None

def clean_duplicate_ots_alternative(conn) -> int:
    """
    Limpiar OTs duplicadas usando método alternativo que evita safe update mode
    Retorna el número de registros eliminados
    """
    max_retries = 3
    base_wait_time = 5
    
    for attempt in range(max_retries):
        try:
            # Verificar conexión antes de proceder
            if not verify_connection(conn):
                logging.warning(f"⚠️ Conexión perdida, reconectando... (intento {attempt + 1})")
                conn = connect_to_mysql()
                if not conn:
                    if attempt < max_retries - 1:
                        time.sleep(base_wait_time * (2 ** attempt))
                        continue
                    else:
                        logging.error("❌ No se pudo reconectar a MySQL")
                        return 0
            
            cursor = conn.cursor()
            
            # Configurar timeout más largo para operaciones grandes
            cursor.execute("SET SESSION wait_timeout = 300")  # 5 minutos
            cursor.execute("SET SESSION interactive_timeout = 300")
            
            # Verificar si existe columna id (PK)
            cursor.execute("SHOW COLUMNS FROM backlog_nacional LIKE 'id'")
            has_id_column = cursor.fetchone() is not None
            
            if not has_id_column:
                # Crear columna id como PK si no existe
                logging.info("🔧 Creando columna id como clave primaria...")
                cursor.execute("ALTER TABLE backlog_nacional ADD COLUMN id INT AUTO_INCREMENT PRIMARY KEY FIRST")
                conn.commit()
                logging.info("✅ Columna id creada exitosamente")
            
            # Contar registros totales antes de limpiar
            cursor.execute("SELECT COUNT(*) FROM backlog_nacional")
            total_registros_antes = cursor.fetchone()[0]
            
            # Contar OTs duplicadas y registros duplicados
            cursor.execute("""
                SELECT OT_LL, COUNT(*) as cnt 
                FROM backlog_nacional 
                GROUP BY OT_LL 
                HAVING cnt > 1
            """)
            ots_duplicadas = cursor.fetchall()
            
            if not ots_duplicadas:
                logging.info("✅ No hay OTs duplicadas")
                cursor.close()
                return 0
            
            # Calcular total de registros que serán duplicados a eliminar
            total_registros_duplicados = sum(cnt - 1 for _, cnt in ots_duplicadas)
            total_ots_duplicadas = len(ots_duplicadas)
            
            logging.info(f"🔍 Encontradas {total_ots_duplicadas} OTs duplicadas con {total_registros_duplicados} registros a eliminar...")
            
            # Método más seguro: Procesar en lotes pequeños si hay muchas OTs duplicadas
            if total_ots_duplicadas > 1000:
                logging.info("📊 Procesando duplicados en lotes pequeños...")
                return clean_duplicates_in_batches(conn, cursor)
            
            # Método alternativo: Crear tabla temporal con los registros únicos
            # PRIORIDAD: 1) OTs gestionadas, 2) Más recientes, 3) Mayor ID
            cursor.execute("""
                CREATE TEMPORARY TABLE temp_backlog_nacional AS
                SELECT * FROM backlog_nacional b1
                WHERE b1.id = (
                    SELECT b2.id 
                    FROM backlog_nacional b2 
                    WHERE b2.OT_LL = b1.OT_LL
                    ORDER BY 
                        -- Prioridad 1: OTs gestionadas primero (ESTADO o DESCRIPCION no vacía)
                        CASE 
                            WHEN b2.ESTADO = 'GESTIONADA' THEN 0
                            WHEN b2.DESCRIPCION IS NOT NULL AND TRIM(b2.DESCRIPCION) <> '' THEN 0
                            WHEN b2.usuario_descripcion IS NOT NULL AND TRIM(b2.usuario_descripcion) <> '' THEN 0
                            ELSE 1 
                        END,
                        -- Prioridad 2: Fecha más reciente
                        CASE 
                            WHEN b2.FECHA_HORA_CREADO IS NOT NULL THEN b2.FECHA_HORA_CREADO 
                            ELSE '1900-01-01 00:00:00' 
                        END DESC,
                        -- Prioridad 3: ID mayor como último criterio
                        b2.id DESC
                    LIMIT 1
                )
            """)
            
            # Contar registros en tabla temporal
            cursor.execute("SELECT COUNT(*) FROM temp_backlog_nacional")
            registros_unicos = cursor.fetchone()[0]
            
            # Limpiar tabla original
            cursor.execute("DELETE FROM backlog_nacional")
            
            # Insertar registros únicos de vuelta
            cursor.execute("""
                INSERT INTO backlog_nacional 
                SELECT * FROM temp_backlog_nacional
            """)
            
            # Eliminar tabla temporal
            cursor.execute("DROP TEMPORARY TABLE temp_backlog_nacional")
            
            # Calcular registros eliminados correctamente
            registros_eliminados = total_registros_antes - registros_unicos
            conn.commit()
            
            logging.info(f"✅ Limpieza completada: {registros_eliminados} registros duplicados eliminados")
            
            # Verificar que no queden duplicados
            cursor.execute("""
                SELECT COUNT(*) as duplicados_restantes 
                FROM (
                    SELECT OT_LL, COUNT(*) as cnt 
                    FROM backlog_nacional 
                    GROUP BY OT_LL 
                    HAVING cnt > 1
                ) as duplicados
            """)
            duplicados_restantes = cursor.fetchone()[0]
            
            if duplicados_restantes > 0:
                logging.warning(f"⚠️ CRÍTICO: Aún quedan {duplicados_restantes} OTs duplicadas después de la limpieza")
                # Intentar una limpieza adicional con DELETE directo
                logging.info("🔄 Intentando limpieza adicional directa...")
                try:
                    # Deshabilitar safe update mode temporalmente
                    cursor.execute("SET SQL_SAFE_UPDATES = 0")
                    
                    # Eliminar duplicados manteniendo el que tiene mayor prioridad
                    # Usar subconsulta para obtener el ID correcto a mantener
                    cursor.execute("""
                        DELETE b1 FROM backlog_nacional b1
                        WHERE b1.id NOT IN (
                            SELECT id FROM (
                                SELECT b2.id 
                                FROM backlog_nacional b2
                                INNER JOIN (
                                    SELECT OT_LL, COUNT(*) as cnt 
                                    FROM backlog_nacional 
                                    GROUP BY OT_LL 
                                    HAVING cnt > 1
                                ) dups ON b2.OT_LL = dups.OT_LL
                                WHERE b2.OT_LL = b1.OT_LL
                                ORDER BY 
                                    CASE 
                                        WHEN b2.ESTADO = 'GESTIONADA' THEN 0
                                        WHEN b2.DESCRIPCION IS NOT NULL AND TRIM(b2.DESCRIPCION) <> '' THEN 0
                                        WHEN b2.usuario_descripcion IS NOT NULL AND TRIM(b2.usuario_descripcion) <> '' THEN 0
                                        ELSE 1 
                                    END,
                                    CASE 
                                        WHEN b2.FECHA_HORA_CREADO IS NOT NULL THEN b2.FECHA_HORA_CREADO 
                                        ELSE '1900-01-01 00:00:00' 
                                    END DESC,
                                    b2.id DESC
                                LIMIT 1
                            ) AS keep_id
                        )
                        AND b1.OT_LL IN (
                            SELECT OT_LL FROM (
                                SELECT OT_LL, COUNT(*) as cnt 
                                FROM backlog_nacional 
                                GROUP BY OT_LL 
                                HAVING cnt > 1
                            ) AS dups_check
                        )
                    """)
                    
                    eliminados_adicionales = cursor.rowcount
                    conn.commit()
                except Exception as e:
                    logging.error(f"❌ Error en limpieza adicional: {e}")
                    eliminados_adicionales = 0
                    conn.rollback()
                finally:
                    # Siempre restaurar safe update mode, incluso si falla el DELETE
                    try:
                        cursor.execute("SET SQL_SAFE_UPDATES = 1")
                    except Exception:
                        pass
                
                if eliminados_adicionales > 0:
                    logging.info(f"✅ Limpieza adicional: {eliminados_adicionales} registros eliminados")
                    
                    # Verificar nuevamente
                    cursor.execute("""
                        SELECT COUNT(*) as duplicados_restantes 
                        FROM (
                            SELECT OT_LL, COUNT(*) as cnt 
                            FROM backlog_nacional 
                            GROUP BY OT_LL 
                            HAVING cnt > 1
                        ) as duplicados
                    """)
                    duplicados_restantes = cursor.fetchone()[0]
                    
                    if duplicados_restantes > 0:
                        logging.error(f"❌ ERROR CRÍTICO: Aún quedan {duplicados_restantes} OTs duplicadas después de limpieza adicional")
                    else:
                        logging.info("✅ Limpieza adicional exitosa: No quedan duplicados")
                        registros_eliminados += eliminados_adicionales
                else:
                    logging.warning("⚠️ Limpieza adicional no eliminó registros")
            else:
                logging.info("✅ Verificación: No quedan OTs duplicadas")
            
            cursor.close()
            return registros_eliminados
            
        except mysql.connector.Error as e:
            error_code = e.errno if hasattr(e, 'errno') else None
            error_msg = str(e).upper()
            
            if error_code == 2013 or 'LOST CONNECTION' in error_msg:
                logging.warning(f"⚠️ Conexión perdida durante limpieza (intento {attempt + 1}): {e}")
                if attempt < max_retries - 1:
                    logging.info(f"🔄 Reintentando en {base_wait_time * (2 ** attempt)} segundos...")
                    time.sleep(base_wait_time * (2 ** attempt))
                    # Intentar reconectar
                    conn = connect_to_mysql()
                    if not conn:
                        continue
                else:
                    logging.error("❌ No se pudo completar la limpieza después de todos los intentos")
                    return 0
            else:
                logging.error(f"❌ Error MySQL limpiando duplicados: {e}")
                if attempt < max_retries - 1:
                    time.sleep(base_wait_time)
                    continue
                else:
                    return 0
                    
        except Exception as e:
            logging.error(f"❌ Error inesperado limpiando duplicados: {e}")
            if attempt < max_retries - 1:
                time.sleep(base_wait_time)
                continue
            else:
                return 0
    
    return 0

def clean_duplicates_in_batches(conn, cursor) -> int:
    """
    Limpiar duplicados procesando en lotes pequeños para evitar timeouts
    Elimina TODAS las filas duplicadas de cada OT, manteniendo solo una según prioridad
    """
    try:
        # Contar registros antes
        cursor.execute("SELECT COUNT(*) FROM backlog_nacional")
        total_registros_antes = cursor.fetchone()[0]
        
        logging.info("📦 Procesando duplicados en lotes de 100 OTs...")
        
        # Obtener OTs duplicadas en lotes
        batch_size = 100
        offset = 0
        total_eliminados = 0
        
        while True:
            # Obtener lote de OTs duplicadas
            cursor.execute("""
                SELECT OT_LL, COUNT(*) as cnt
                FROM backlog_nacional 
                GROUP BY OT_LL 
                HAVING cnt > 1
                ORDER BY OT_LL
                LIMIT %s OFFSET %s
            """, (batch_size, offset))
            
            batch_ots = cursor.fetchall()
            if not batch_ots:
                break
            
            logging.info(f"🔍 Procesando lote {offset//batch_size + 1}: {len(batch_ots)} OTs")
            
            # Procesar cada OT del lote
            for ot, count in batch_ots:
                try:
                    # Obtener TODOS los registros de esta OT con información completa
                    # PRIORIDAD: 1) OTs gestionadas, 2) Más recientes, 3) Mayor ID
                    cursor.execute("""
                        SELECT id, ESTADO, DESCRIPCION, usuario_descripcion, FECHA_HORA_CREADO
                        FROM backlog_nacional 
                        WHERE OT_LL = %s
                        ORDER BY 
                            -- Prioridad 1: OTs gestionadas primero (ESTADO o DESCRIPCION no vacía)
                            CASE 
                                WHEN ESTADO = 'GESTIONADA' THEN 0
                                WHEN DESCRIPCION IS NOT NULL AND TRIM(DESCRIPCION) <> '' THEN 0
                                WHEN usuario_descripcion IS NOT NULL AND TRIM(usuario_descripcion) <> '' THEN 0
                                ELSE 1 
                            END,
                            -- Prioridad 2: Fecha más reciente
                            CASE 
                                WHEN FECHA_HORA_CREADO IS NOT NULL THEN FECHA_HORA_CREADO 
                                ELSE '1900-01-01 00:00:00' 
                            END DESC,
                            -- Prioridad 3: ID mayor como último criterio
                            id DESC
                    """, (ot,))
                    
                    registros = cursor.fetchall()
                    if len(registros) <= 1:
                        continue
                    
                    # Mantener solo el primero (según prioridad) y eliminar TODOS los demás
                    id_a_mantener = registros[0][0]  # Primer registro según prioridad
                    ids_a_eliminar = [reg[0] for reg in registros[1:]]  # Todos los demás
                    
                    if ids_a_eliminar and len(ids_a_eliminar) > 0:
                        # Verificar que no estamos eliminando el ID que queremos mantener
                        if id_a_mantener in ids_a_eliminar:
                            ids_a_eliminar = [id_val for id_val in ids_a_eliminar if id_val != id_a_mantener]
                        
                        if ids_a_eliminar:
                            placeholders = ','.join(['%s'] * len(ids_a_eliminar))
                            cursor.execute(f"""
                                DELETE FROM backlog_nacional 
                                WHERE id IN ({placeholders})
                                AND OT_LL = %s
                            """, ids_a_eliminar + [ot])
                            
                            eliminados = cursor.rowcount
                            total_eliminados += eliminados
                            
                            if eliminados != len(ids_a_eliminar):
                                logging.warning(f"⚠️ OT {ot}: Se esperaba eliminar {len(ids_a_eliminar)} pero se eliminaron {eliminados}")
                            
                            logging.info(f"🗑️ OT {ot}: eliminados {eliminados} duplicados de {count} registros totales (mantenido ID {id_a_mantener})")
                    
                except Exception as e:
                    logging.warning(f"⚠️ Error procesando OT {ot}: {e}")
                    continue
            
            # Commit del lote
            conn.commit()
            offset += batch_size
            
            # Pequeña pausa entre lotes
            time.sleep(0.5)
        
        # Verificar que realmente se eliminaron
        cursor.execute("SELECT COUNT(*) FROM backlog_nacional")
        total_registros_despues = cursor.fetchone()[0]
        registros_eliminados_real = total_registros_antes - total_registros_despues
        
        # Verificar que no queden duplicados
        cursor.execute("""
            SELECT COUNT(*) as duplicados_restantes 
            FROM (
                SELECT OT_LL, COUNT(*) as cnt 
                FROM backlog_nacional 
                GROUP BY OT_LL 
                HAVING cnt > 1
            ) as duplicados
        """)
        duplicados_restantes = cursor.fetchone()[0]
        
        if duplicados_restantes > 0:
            logging.warning(f"⚠️ CRÍTICO: Aún quedan {duplicados_restantes} OTs duplicadas después de la limpieza")
            # Intentar una limpieza adicional
            logging.info("🔄 Intentando limpieza adicional...")
            cursor.execute("""
                DELETE b1 FROM backlog_nacional b1
                INNER JOIN backlog_nacional b2 
                WHERE b1.OT_LL = b2.OT_LL 
                AND b1.id < b2.id
                AND NOT (
                    b1.ESTADO = 'GESTIONADA' 
                    OR (b1.DESCRIPCION IS NOT NULL AND TRIM(b1.DESCRIPCION) <> '')
                    OR (b1.usuario_descripcion IS NOT NULL AND TRIM(b1.usuario_descripcion) <> '')
                )
            """)
            conn.commit()
            
            # Verificar nuevamente
            cursor.execute("""
                SELECT COUNT(*) as duplicados_restantes 
                FROM (
                    SELECT OT_LL, COUNT(*) as cnt 
                    FROM backlog_nacional 
                    GROUP BY OT_LL 
                    HAVING cnt > 1
                ) as duplicados
            """)
            duplicados_restantes = cursor.fetchone()[0]
            
            if duplicados_restantes > 0:
                logging.error(f"❌ ERROR: Aún quedan {duplicados_restantes} OTs duplicadas después de limpieza adicional")
            else:
                logging.info("✅ Limpieza adicional exitosa: No quedan duplicados")
        else:
            logging.info("✅ Verificación: No quedan OTs duplicadas")
        
        logging.info(f"✅ Limpieza por lotes completada: {registros_eliminados_real} registros eliminados (antes: {total_registros_antes}, después: {total_registros_despues})")
        return registros_eliminados_real
        
    except Exception as e:
        logging.error(f"❌ Error en limpieza por lotes: {e}")
        conn.rollback()
        return 0

def clean_duplicate_ots(conn) -> int:
    """
    Limpiar OTs duplicadas manteniendo el registro más reciente
    Retorna el número de registros eliminados
    """
    try:
        cursor = conn.cursor()
        
        # Verificar si existe columna id (PK)
        cursor.execute("SHOW COLUMNS FROM backlog_nacional LIKE 'id'")
        has_id_column = cursor.fetchone() is not None
        
        if not has_id_column:
            # Crear columna id como PK si no existe
            logging.info("🔧 Creando columna id como clave primaria...")
            cursor.execute("ALTER TABLE backlog_nacional ADD COLUMN id INT AUTO_INCREMENT PRIMARY KEY FIRST")
            conn.commit()
            logging.info("✅ Columna id creada exitosamente")
        
        # Contar duplicados antes de limpiar
        cursor.execute("""
            SELECT COUNT(*) as total_duplicados 
            FROM (
                SELECT OT_LL, COUNT(*) as cnt 
                FROM backlog_nacional 
                GROUP BY OT_LL 
                HAVING cnt > 1
            ) as duplicados
        """)
        total_duplicados = cursor.fetchone()[0]
        
        if total_duplicados == 0:
            logging.info("✅ No hay OTs duplicadas")
            cursor.close()
            return 0
        
        logging.info(f"🔍 Encontradas {total_duplicados} OTs duplicadas, limpiando...")
        
        # Deshabilitar safe update mode temporalmente
        cursor.execute("SET SQL_SAFE_UPDATES = 0")
        
        # Eliminar duplicados PROTEGIENDO OTs gestionadas
        # PRIORIDAD: 1) OTs gestionadas, 2) Más recientes, 3) Mayor ID
        cursor.execute("""
            DELETE b1 FROM backlog_nacional b1
            INNER JOIN backlog_nacional b2 
            WHERE b1.OT_LL = b2.OT_LL 
            AND b1.id != (
                SELECT sub.id FROM (
                    SELECT b3.id
                    FROM backlog_nacional b3
                    WHERE b3.OT_LL = b1.OT_LL
                    ORDER BY 
                        -- Prioridad 1: OTs gestionadas primero
                        CASE WHEN b3.ESTADO = 'GESTIONADA' OR b3.DESCRIPCION IS NOT NULL THEN 0 ELSE 1 END,
                        -- Prioridad 2: Fecha más reciente
                        CASE WHEN b3.FECHA_HORA_CREADO IS NOT NULL THEN b3.FECHA_HORA_CREADO ELSE '1900-01-01' END DESC,
                        -- Prioridad 3: ID mayor como último criterio
                        b3.id DESC
                    LIMIT 1
                ) sub
            )
        """)
        
        # Rehabilitar safe update mode
        cursor.execute("SET SQL_SAFE_UPDATES = 1")
        
        registros_eliminados = cursor.rowcount
        conn.commit()
        
        logging.info(f"✅ Limpieza completada: {registros_eliminados} registros duplicados eliminados")
        
        # Verificar que no queden duplicados
        cursor.execute("""
            SELECT COUNT(*) as duplicados_restantes 
            FROM (
                SELECT OT_LL, COUNT(*) as cnt 
                FROM backlog_nacional 
                GROUP BY OT_LL 
                HAVING cnt > 1
            ) as duplicados
        """)
        duplicados_restantes = cursor.fetchone()[0]
        
        if duplicados_restantes > 0:
            logging.warning(f"⚠️ Aún quedan {duplicados_restantes} OTs duplicadas")
        else:
            logging.info("✅ Verificación: No quedan OTs duplicadas")
        
        cursor.close()
        return registros_eliminados
        
    except Exception as e:
        logging.error(f"❌ Error limpiando duplicados: {e}")
        conn.rollback()
        return 0

def insert_data_to_mysql(conn, df: pl.DataFrame) -> bool:
    """
    Insertar datos en MySQL preservando gestiones comerciales
    VERSIÓN MEJORADA: SIN MEMORIA - Todo se maneja en MySQL con tablas temporales
    
    Ventajas:
    - ✅ No usa memoria RAM (no se pierden datos si el bot se cae)
    - ✅ Transacciones atómicas (todo o nada)
    - ✅ Más rápido (operaciones SQL nativas)
    - ✅ Más seguro (no hay riesgo de pérdida de datos)
    """
    try:
        cursor = conn.cursor()
        
        # Obtener columnas de la tabla
        cursor.execute("SHOW COLUMNS FROM backlog_nacional")
        table_columns = [row[0].lower() for row in cursor.fetchall()]
        
        # Filtrar DataFrame a columnas existentes
        valid_columns = [col for col in df.columns if col in table_columns]
        if not valid_columns:
            logging.warning("⚠️ No hay columnas válidas para insertar")
            cursor.close()
            return False
        
        df_filtered = df.select(valid_columns)
        
        # Configurar timeouts de sesión
        cursor.execute("SET SESSION innodb_lock_wait_timeout = 120")
        cursor.execute("SET SESSION lock_wait_timeout = 120")
        cursor.execute("SET SESSION wait_timeout = 300")
        
        logging.info("🔄 INICIANDO PROCESO SIN MEMORIA - Todo en MySQL")
        
        # ========== PASO 1: CREAR TABLA TEMPORAL PARA NUEVOS DATOS ==========
        logging.info("1️⃣ Creando tabla temporal para nuevos datos...")
        
        # Crear tabla temporal con la misma estructura que backlog_nacional
        try:
            cursor.execute("DROP TEMPORARY TABLE IF EXISTS temp_nuevos_datos")
        except:
            pass  # Ignorar si no existe
            
        cursor.execute("""
            CREATE TEMPORARY TABLE temp_nuevos_datos AS 
            SELECT * FROM backlog_nacional WHERE 1=0
        """)
        
        # Insertar nuevos datos en tabla temporal
        columns = ', '.join(df_filtered.columns)
        placeholders = ', '.join(['%s'] * len(df_filtered.columns))
        sql = f"INSERT INTO temp_nuevos_datos ({columns}) VALUES ({placeholders})"
        
        data = [tuple(row) for row in df_filtered.to_numpy()]
        
        # Insertar en lotes
        batch_size = 2000
        total_inserted = 0
        
        for i in range(0, len(data), batch_size):
            batch = data[i:i + batch_size]
            cursor.executemany(sql, batch)
            total_inserted += len(batch)
            logging.info(f"📊 Lote temporal: {len(batch)} registros (Total: {total_inserted}/{len(data)})")
        
        conn.commit()
        logging.info(f"✅ Nuevos datos cargados en tabla temporal: {total_inserted} registros")
        
        # ========== PASO 2: CREAR TABLA TEMPORAL PARA GESTIONES A PRESERVAR ==========
        logging.info("2️⃣ Identificando gestiones a preservar...")
        
        try:
            cursor.execute("DROP TEMPORARY TABLE IF EXISTS temp_gestiones_preservar")
        except:
            pass  # Ignorar si no existe
        
        # Crear tabla temporal para gestiones que deben preservarse
        cursor.execute("""
             CREATE TEMPORARY TABLE temp_gestiones_preservar AS
             SELECT 
                 b.OT,
                 b.DESCRIPCION,
                 b.usuario_descripcion,
                 b.fecha_descripcion,
                 b.solicitud,
                 b.motivo,
                 b.fecha_reagendar,
                 b.franja_horaria
             FROM backlog_nacional b
             INNER JOIN temp_nuevos_datos n ON b.OT = n.OT
             WHERE (b.DESCRIPCION IS NOT NULL OR b.usuario_descripcion IS NOT NULL)
         """)
        
        # Contar gestiones a preservar
        cursor.execute("SELECT COUNT(*) FROM temp_gestiones_preservar")
        gestiones_count = cursor.fetchone()[0]
        logging.info(f"💾 Gestiones a preservar: {gestiones_count}")
        
        # ========== PASO 3: ARCHIVAR GESTIONES QUE YA NO ESTÁN EN EL EXCEL ==========
        logging.info("3️⃣ Archivando gestiones que ya no están en el Excel...")
        
        # Archivar gestiones que NO están en los nuevos datos
        cursor.execute("""
             INSERT INTO backlog_nacional_historial 
             (CUENTA, OT, DESCRIPCION, usuario_descripcion, fecha_descripcion, 
              fecha_archivado, solicitud, motivo, fecha_reagendar, franja_horaria, razon_comercial)
             SELECT 
                 CAST(b.CUENTA AS CHAR) as CUENTA,
                 CAST(b.OT AS CHAR) as OT,
                 b.DESCRIPCION,
                 b.usuario_descripcion,
                 b.fecha_descripcion,
                 NOW() as fecha_archivado,
                 b.solicitud,
                 b.motivo,
                 b.fecha_reagendar,
                 b.franja_horaria,
                 COALESCE(NULLIF(b.RAZON_WF, ''), NULLIF(b.RAZON_RR, ''), b.COD_RAZON_RR) as razon_comercial
             FROM backlog_nacional b
             LEFT JOIN temp_nuevos_datos n ON b.OT = n.OT
             WHERE (b.DESCRIPCION IS NOT NULL OR b.usuario_descripcion IS NOT NULL)
             AND n.OT IS NULL
         """)
        
        archivados = cursor.rowcount
        if archivados > 0:
            logging.info(f"📦 Archivados {archivados} registros gestionados que ya no están en el Excel")
        
        # ========== PASO 4: REEMPLAZAR DATOS PRINCIPALES ==========
        logging.info("4️⃣ Reemplazando datos principales...")
        
        # Limpiar tabla principal
        cursor.execute("DELETE FROM backlog_nacional")
        logging.info("🗑️ Tabla principal limpiada")
        
        # Insertar nuevos datos desde tabla temporal
        cursor.execute("""
            INSERT INTO backlog_nacional 
            SELECT * FROM temp_nuevos_datos
        """)
        
        nuevos_registros = cursor.rowcount
        logging.info(f"📊 Nuevos datos insertados: {nuevos_registros} registros")
        
        # ========== PASO 5: RESTAURAR GESTIONES PRESERVADAS ==========
        logging.info("5️⃣ Restaurando gestiones preservadas...")
        
        # Restaurar gestiones usando JOIN (más eficiente que loops)
        cursor.execute("""
             UPDATE backlog_nacional b
             INNER JOIN temp_gestiones_preservar g ON b.OT = g.OT
             SET 
                 b.DESCRIPCION = g.DESCRIPCION,
                 b.usuario_descripcion = g.usuario_descripcion,
                 b.fecha_descripcion = g.fecha_descripcion,
                 b.solicitud = g.solicitud,
                 b.motivo = g.motivo,
                 b.fecha_reagendar = g.fecha_reagendar,
                 b.franja_horaria = g.franja_horaria,
                 b.ESTADO = 'GESTIONADA'
         """)
        
        restaurados = cursor.rowcount
        if restaurados > 0:
            logging.info(f"✅ Restauradas {restaurados} gestiones en registros que siguen en el Excel")
        
        # ========== PASO 6: LIMPIAR DUPLICADOS ==========
        logging.info("6️⃣ Limpiando duplicados...")
        
        # Usar función existente de limpieza
        registros_eliminados = clean_duplicate_ots_alternative(conn)
        
        if registros_eliminados > 0:
            logging.info(f"🧹 Limpieza de duplicados: {registros_eliminados} registros eliminados")
        
        # ========== COMMIT FINAL ==========
        conn.commit()
        
        # ========== RESUMEN ==========
        logging.info(f"""
📊 RESUMEN DEL PROCESO SIN MEMORIA:
   • Nuevos registros: {nuevos_registros:,}
   • Gestiones archivadas: {archivados}
   • Gestiones restauradas: {restaurados}
   • Duplicados eliminados: {registros_eliminados}
   • ✅ Proceso completado sin usar memoria RAM
        """)
        
        cursor.close()
        return True
        
    except Exception as e:
        logging.error(f"❌ Error en proceso sin memoria: {e}")
        conn.rollback()
        return False

def get_system_status() -> str:
    """Obtener información del estado del sistema"""
    global last_update_time, script_start_time, last_processed_filename, consecutive_errors
    
    current_time = now_colombia()
    uptime = current_time - script_start_time
    
    status_info = []
    status_info.append("=" * 60)
    status_info.append("📊 ESTADO DEL SISTEMA")
    status_info.append("=" * 60)
    status_info.append(f"🕐 Hora actual: {current_time.strftime('%Y-%m-%d %H:%M:%S')}")
    status_info.append(f"🚀 Iniciado: {script_start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    status_info.append(f"⏱️ Tiempo activo: {str(uptime).split('.')[0]}")
    
    if last_update_time:
        time_since_update = current_time - last_update_time
        status_info.append(f"✅ Última actualización: {last_update_time.strftime('%Y-%m-%d %H:%M:%S')}")
        status_info.append(f"⏳ Tiempo desde última actualización: {str(time_since_update).split('.')[0]}")
    else:
        status_info.append("❌ Sin actualizaciones desde el inicio")
    
    if last_processed_filename:
        status_info.append(f"📄 Último archivo procesado: {last_processed_filename}")
    else:
        status_info.append("📄 Ningún archivo procesado aún")
    
    status_info.append(f"⚠️ Errores consecutivos: {consecutive_errors}/{max_consecutive_errors}")
    status_info.append("=" * 60)
    
    return "\n".join(status_info)

def print_system_status():
    """Imprimir estado del sistema en consola y log"""
    status = get_system_status()
    print(status)
    for line in status.split('\n'):
        if line.strip():
            logging.info(line)

def get_system_status_json() -> Dict[str, Any]:
    """Obtener información del estado del sistema en formato JSON"""
    global last_update_time, script_start_time, last_processed_filename, consecutive_errors
    global last_file_date, total_files_processed, last_error_message
    
    current_time = now_colombia()
    uptime = current_time - script_start_time
    
    status_data = {
        "timestamp": current_time.isoformat(),
        "uptime_seconds": int(uptime.total_seconds()),
        "uptime_formatted": str(uptime).split('.')[0],
        "script_start_time": script_start_time.isoformat(),
        "last_update_time": last_update_time.isoformat() if last_update_time else None,
        "time_since_last_update_seconds": int((current_time - last_update_time).total_seconds()) if last_update_time else None,
        "last_processed_filename": last_processed_filename,
        "last_file_date": last_file_date.isoformat() if last_file_date else None,
        "total_files_processed": total_files_processed,
        "consecutive_errors": consecutive_errors,
        "max_consecutive_errors": max_consecutive_errors,
        "last_error_message": last_error_message,
        "status": "running" if consecutive_errors < max_consecutive_errors else "error",
        "bot_type": "sharepoint_monitor",
        "service": "Bot OTC SharePoint Monitor"
    }
    
    return status_data

# Endpoints Flask
@app.route('/status', methods=['GET'])
def status_endpoint():
    """Endpoint para obtener el estado del sistema en JSON"""
    try:
        status_data = get_system_status_json()
        return jsonify(status_data), 200
    except Exception as e:
        return jsonify({
            "error": "Error obteniendo estado del sistema",
            "message": str(e),
            "timestamp": now_colombia().isoformat()
        }), 500

@app.route('/health', methods=['GET'])
def health_endpoint():
    """Endpoint simple de health check"""
    return jsonify({
        "status": "healthy",
        "timestamp": now_colombia().isoformat(),
        "service": "SharePoint Monitor Service"
    }), 200


@app.route('/update-cookies', methods=['POST'])
def update_cookies_endpoint():
    """
    Endpoint para actualizar cookies de SharePoint remotamente
    Enviar JSON: {"FedAuth": "valor", "rtFa": "valor"}
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                "status": "error",
                "message": "Se requiere JSON con FedAuth y rtFa"
            }), 400
        
        fed_auth = data.get('FedAuth')
        rt_fa = data.get('rtFa')
        
        if not fed_auth or not rt_fa:
            return jsonify({
                "status": "error",
                "message": "Se requieren ambos campos: FedAuth y rtFa"
            }), 400
        
        # Guardar nuevas cookies
        new_cookies = {
            "FedAuth": fed_auth,
            "rtFa": rt_fa
        }
        
        with open(COOKIES_FILE, 'w', encoding='utf-8') as f:
            json.dump(new_cookies, f, indent=2)
        
        logging.info("🔑 Cookies de SharePoint actualizadas remotamente")
        
        # Verificar que las cookies funcionan
        session = create_sharepoint_session()
        if session:
            test_url = f"{SHAREPOINT_BASE}/sites/AseguramientoOperacionCGO/_api/web"
            try:
                resp = session.get(test_url, timeout=10)
                if resp.status_code == 200:
                    return jsonify({
                        "status": "success",
                        "message": "Cookies actualizadas y verificadas correctamente",
                        "timestamp": now_colombia().isoformat()
                    }), 200
                elif resp.status_code == 401:
                    return jsonify({
                        "status": "warning",
                        "message": "Cookies guardadas pero parecen inválidas (401)",
                        "timestamp": now_colombia().isoformat()
                    }), 200
            except Exception as e:
                logging.warning(f"⚠️ No se pudo verificar cookies: {e}")
        
        return jsonify({
            "status": "success",
            "message": "Cookies actualizadas (no se pudo verificar conexión)",
            "timestamp": now_colombia().isoformat()
        }), 200
        
    except Exception as e:
        logging.error(f"❌ Error actualizando cookies: {e}")
        return jsonify({
            "status": "error",
            "message": str(e),
            "timestamp": now_colombia().isoformat()
        }), 500


@app.route('/cookies-status', methods=['GET'])
def cookies_status_endpoint():
    """Verificar estado de las cookies de SharePoint"""
    try:
        # Verificar si existe el archivo
        if not os.path.exists(COOKIES_FILE):
            return jsonify({
                "status": "error",
                "message": "No existe archivo cookies.json",
                "valid": False
            }), 200
        
        # Intentar conectar a SharePoint
        session = create_sharepoint_session()
        if not session:
            return jsonify({
                "status": "error",
                "message": "No se pudo crear sesión con las cookies actuales",
                "valid": False
            }), 200
        
        # Probar conexión
        test_url = f"{SHAREPOINT_BASE}/sites/AseguramientoOperacionCGO/_api/web"
        try:
            resp = session.get(test_url, timeout=10)
            if resp.status_code == 200:
                return jsonify({
                    "status": "success",
                    "message": "Cookies válidas y funcionando",
                    "valid": True,
                    "timestamp": now_colombia().isoformat()
                }), 200
            elif resp.status_code == 401:
                return jsonify({
                    "status": "expired",
                    "message": "Cookies expiradas - necesitan actualización",
                    "valid": False,
                    "timestamp": now_colombia().isoformat()
                }), 200
            else:
                return jsonify({
                    "status": "unknown",
                    "message": f"Respuesta inesperada: {resp.status_code}",
                    "valid": False
                }), 200
        except Exception as e:
            return jsonify({
                "status": "error",
                "message": f"Error de conexión: {str(e)}",
                "valid": False
            }), 200
            
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500


@app.route('/duplicates', methods=['GET'])
def duplicates_endpoint():
    """Endpoint para verificar y limpiar duplicados"""
    try:
        conn = connect_to_mysql()
        if not conn:
            return jsonify({
                "error": "No se pudo conectar a la base de datos",
                "timestamp": now_colombia().isoformat()
            }), 500
        
        cursor = conn.cursor()
        
        # Verificar duplicados
        cursor.execute("""
            SELECT OT_LL, COUNT(*) as cantidad_duplicados
            FROM backlog_nacional 
            GROUP BY OT_LL 
            HAVING COUNT(*) > 1 
            ORDER BY cantidad_duplicados DESC
            LIMIT 20
        """)
        duplicates = cursor.fetchall()
        
        # Contar total de duplicados
        cursor.execute("""
            SELECT COUNT(*) as total_duplicados 
            FROM (
                SELECT OT_LL, COUNT(*) as cnt 
                FROM backlog_nacional 
                GROUP BY OT_LL 
                HAVING cnt > 1
            ) as duplicados
        """)
        total_duplicados = cursor.fetchone()[0]
        
        # Limpiar duplicados si se solicita
        cleaned_count = 0
        if request.args.get('clean') == 'true':
            cleaned_count = clean_duplicate_ots_alternative(conn)
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "total_duplicates": total_duplicados,
            "duplicates_list": [{"ot_ll": ot, "count": count} for ot, count in duplicates],
            "cleaned_count": cleaned_count,
            "timestamp": now_colombia().isoformat()
        }), 200
        
    except Exception as e:
        return jsonify({
            "error": "Error verificando duplicados",
            "message": str(e),
            "timestamp": now_colombia().isoformat()
        }), 500



_flask_started = False

def start_flask_server():
    """Iniciar servidor Flask en un hilo separado"""
    try:
        logging.info(f"🌐 Iniciando servidor Flask en puerto {flask_port}")
        app.run(host='0.0.0.0', port=flask_port, debug=False, use_reloader=False)
    except Exception as e:
        logging.error(f"❌ Error iniciando servidor Flask: {e}")

def check_for_new_sharepoint_file() -> bool:
    """Verificar si hay un nuevo archivo en SharePoint y procesarlo"""
    global last_processed_filename, consecutive_errors, last_update_time
    global last_file_date, total_files_processed, last_error_message

    ensure_download_folder()

    session = create_sharepoint_session()
    if not session:
        last_error_message = "No se pudo crear sesión de SharePoint"
        consecutive_errors += 1
        return False

    try:
        latest_file = get_latest_sharepoint_file(session)
        if not latest_file:
            logging.info("📭 No hay archivos nuevos en SharePoint")
            consecutive_errors = 0
            return True

        filename = latest_file.get("Name")
        if not filename:
            logging.warning("⚠️ SharePoint devolvió un archivo sin nombre")
            consecutive_errors += 1
            return False

        file_timestamp = parse_sharepoint_timestamp(latest_file.get("TimeLastModified"))
        state = load_state()

        if state.get("last_processed") == filename or last_processed_filename == filename:
            logging.info(f"⏩ Archivo ya procesado: {filename}")
            consecutive_errors = 0
            return True

        logging.info(f"📄 Nuevo archivo en SharePoint: {filename}")
        if file_timestamp:
            logging.info(f"🕒 TimeLastModified: {file_timestamp.strftime('%Y-%m-%d %H:%M:%S')}")

        download_path = download_sharepoint_file(session, latest_file)
        if not download_path:
            last_error_message = "No se pudo descargar archivo desde SharePoint"
            consecutive_errors += 1
            return False

        processed = False
        file_ext = os.path.splitext(filename)[1].lower()
        if file_ext == ".csv":
            processed = process_backlog_file(download_path)
        else:
            processed = process_downloaded_file(download_path, filename)

        if processed:
            last_processed_filename = filename
            last_update_time = now_colombia()
            last_file_date = file_timestamp or last_update_time
            total_files_processed += 1
            last_error_message = None
            save_state(filename, processed_filename=filename)
            consecutive_errors = 0
            logging.info(f"🎉 Procesamiento completado exitosamente - Última actualización: {last_update_time.strftime('%Y-%m-%d %H:%M:%S')}")
            return True

        last_error_message = "Error procesando archivo de SharePoint"
        consecutive_errors += 1
        return False
    except Exception as exc:
        last_error_message = f"Error en verificación de SharePoint: {exc}"
        logging.error(f"❌ {last_error_message}")
        consecutive_errors += 1
        return False
    finally:
        try:
            session.close()
        except Exception:
            pass

def main():
    """Servicio principal de monitoreo 24/7"""
    global consecutive_errors
    
    logging.info("🚀 Iniciando servicio de monitoreo 24/7")
    logging.info("⏰ Verificando cada 5 minutos")
    logging.info(f"☁️ SharePoint: {SHAREPOINT_ROOT}")
    logging.info("💾 Base de datos: pruebas.backlog_nacional")
    logging.info(f"📁 Carpeta de descargas: {DOWNLOAD_DIR}")
    logging.info(f"🌐 API JSON disponible en: http://localhost:{flask_port}/status")
    logging.info("🔧 Endpoints disponibles: /status, /health, /duplicates")
    logging.info("=" * 60)
    
    global _flask_started
    # Iniciar servidor Flask en hilo separado solo una vez
    if not _flask_started:
        flask_thread = threading.Thread(target=start_flask_server, daemon=True)
        flask_thread.start()
        _flask_started = True
        logging.info("✅ Servidor Flask iniciado en hilo separado")
    
    # Mostrar estado inicial del sistema
    print_system_status()
    
    # Verificación inicial
    logging.info("🔄 Verificación inicial al arrancar...")
    check_for_new_sharepoint_file()
    
    # Contador para mostrar estado completo periódicamente
    status_check_counter = 0
    status_check_interval = 12  # Mostrar estado completo cada 12 verificaciones (1 hora aprox)
    
    while True:
        try:
            # Calcular tiempo de espera dinámico
            if consecutive_errors >= max_consecutive_errors:
                wait_minutes = 15  # Esperar más si hay muchos errores
                logging.warning(f"⚠️ Muchos errores consecutivos ({consecutive_errors}), esperando {wait_minutes} minutos")
            else:
                wait_minutes = 5  # Verificación normal cada 5 minutos
            
            # Mostrar próxima verificación y estado
            next_check = (now_colombia() + timedelta(minutes=wait_minutes)).strftime('%H:%M:%S')
            if last_update_time:
                time_since_update = now_colombia() - last_update_time
                logging.info(f"⏳ Próxima verificación: {next_check} ({wait_minutes} min) | Última actualización: {str(time_since_update).split('.')[0]} atrás")
            else:
                logging.info(f"⏳ Próxima verificación: {next_check} ({wait_minutes} min) | Sin actualizaciones aún")
            
            # Esperar
            time.sleep(wait_minutes * 60)
            
            # Verificar salud del bot antes de procesar archivos
            bot_healthy = check_bot_health()
            if not bot_healthy:
                logging.warning("🚨 Bot marcado como no saludable, continuando con verificación...")
            
            # Verificar SharePoint
            logging.info(f"🔍 Verificando SharePoint... [{now_colombia().strftime('%H:%M:%S')}]")
            success = check_for_new_sharepoint_file()
            
            # Incrementar contador y mostrar estado completo si es necesario
            status_check_counter += 1
            if status_check_counter >= status_check_interval:
                logging.info("📊 Mostrando estado completo del sistema:")
                print_system_status()
                status_check_counter = 0  # Reset contador
            
            # Reset contador si hay demasiados errores consecutivos y tuvimos éxito
            if success and consecutive_errors >= max_consecutive_errors:
                consecutive_errors = 0
                logging.info("✅ Errores consecutivos reseteados")
                
        except KeyboardInterrupt:
            logging.info("\n🛑 Servicio detenido por el usuario")
            break
            
        except Exception as e:
            logging.error(f"❌ Error crítico en servicio principal: {e}")
            consecutive_errors += 1
            
            # Si hay demasiados errores, esperar más tiempo
            if consecutive_errors >= max_consecutive_errors:
                logging.error(f"🚨 Demasiados errores consecutivos ({consecutive_errors})")
                logging.info("⏸️ Pausando servicio por 30 minutos...")
                time.sleep(1800)  # 30 minutos
                consecutive_errors = 0  # Reset después de la pausa larga

def run_with_high_availability():
    """
    Sistema de alta disponibilidad que NUNCA se cae
    - Reinicio automático en caso de errores críticos
    - Manejo robusto de todas las excepciones
    - Monitoreo continuo del estado del sistema
    """
    restart_count = 0
    max_restarts_per_hour = 10
    restart_times = []
    
    while True:
        try:
            logging.info("🚀 Iniciando servicio con sistema de alta disponibilidad")
            logging.info(f"📊 Reinicio #{restart_count + 1}")
            
            # Limpiar reintentos antiguos (más de 1 hora)
            current_time = time.time()
            restart_times = [t for t in restart_times if current_time - t < 3600]
            
            # Verificar si hay demasiados reinicios en la última hora
            if len(restart_times) >= max_restarts_per_hour:
                logging.warning(f"⚠️ Demasiados reinicios ({len(restart_times)}) en la última hora")
                logging.info("⏸️ Pausando 10 minutos antes del siguiente intento...")
                time.sleep(600)  # 10 minutos
                restart_times = []  # Reset contador
            
            # Registrar este intento de reinicio
            restart_times.append(current_time)
            restart_count += 1
            
            # Ejecutar el servicio principal
            main()
            
        except KeyboardInterrupt:
            logging.info("\n🛑 Servicio detenido manualmente por el usuario")
            logging.info("✅ Apagado limpio del sistema")
            break
            
        except SystemExit:
            logging.info("🔄 Reinicio solicitado por el sistema")
            continue
            
        except MemoryError:
            logging.error("🚨 ERROR CRÍTICO: Sin memoria disponible")
            logging.info("🔄 Liberando memoria y reiniciando en 60 segundos...")
            import gc
            gc.collect()  # Forzar recolección de basura
            time.sleep(60)
            continue
            
        except OSError as e:
            logging.error(f"🚨 ERROR CRÍTICO del sistema operativo: {e}")
            logging.info("🔄 Reiniciando en 30 segundos...")
            time.sleep(30)
            continue
            
        except Exception as e:
            logging.error(f"🚨 ERROR CRÍTICO INESPERADO: {type(e).__name__}: {e}")
            logging.error(f"📍 Detalles del error: {str(e)}")
            
            # Log del stack trace para debugging
            import traceback
            logging.error(f"📋 Stack trace completo:\n{traceback.format_exc()}")
            
            # Esperar antes de reiniciar (escalamiento exponencial)
            wait_time = min(60 * (2 ** min(len(restart_times), 5)), 300)  # Máximo 5 minutos
            logging.info(f"🔄 Reiniciando automáticamente en {wait_time} segundos...")
            time.sleep(wait_time)
            continue
        
        # Si llegamos aquí, el servicio terminó normalmente (no debería pasar)
        logging.warning("⚠️ El servicio terminó inesperadamente sin error")
        logging.info("🔄 Reiniciando automáticamente en 10 segundos...")
        time.sleep(10)

def create_system_monitor():
    """
    Monitor del sistema que verifica la salud del servicio
    """
    def monitor_thread():
        while True:
            try:
                # Verificar que el servicio esté respondiendo
                time.sleep(300)  # Verificar cada 5 minutos
                
                # Aquí puedes agregar verificaciones adicionales
                # como verificar que Flask esté respondiendo, etc.
                
            except Exception as e:
                logging.error(f"❌ Error en monitor del sistema: {e}")
                
    import threading
    monitor = threading.Thread(target=monitor_thread, daemon=True)
    monitor.start()
    logging.info("🔍 Monitor del sistema iniciado")

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# Configuración de Email
EMAIL_CONFIG = {
    'smtp_server': 'smtp.gmail.com',
    'smtp_port': 587,
    'sender_email': 'backlog326@gmail.com',
    'app_password': 'kzpf obmb yyqf ynts',
    'recipients': [
        'martha.rios@claro.com.co',
        'jaime.chamorro@claro.com.co',
        'Jose.Munoz@claro.com.co',
        'angie.fontecha@claro.com.co',
        'diana.aristizabalg@claro.com.co',
        'JOSEFINA.RIOS@claro.com.co'
    ]
}


def get_region_occidente_data():
    """
    Obtener datos filtrados por REGION OCCIDENTE desde la base de datos
    Equivalente a: Table.SelectRows(Origen, each [REGION] = "REGION OCCIDENTE")
    Retorna un DataFrame de pandas para mejor compatibilidad con tipos mixtos
    """
    import pandas as pd
    
    # Columnas a excluir del reporte
    COLUMNAS_EXCLUIR = [
        'id',
        'descripcion',
        'usuario_descripcion',
        'fecha_descripcion',
        'solicitud',
        'motivo',
        'fecha_reagendar',
        'franja_horaria',
        'ESTADO',
        'FECHA_HORA_CREADO'
    ]
    
    conn = None
    try:
        conn = open_backlog_db_connection()
        if not conn:
            logging.error("❌ No se pudo conectar a la base de datos para obtener datos de REGION OCCIDENTE")
            return None
        
        # Usar pandas para leer directamente desde SQL (maneja mejor tipos mixtos)
        query = """
            SELECT * FROM backlog_nacional 
            WHERE REGION = 'REGION OCCIDENTE'
        """
        
        df = pd.read_sql(query, conn)
        
        if df.empty:
            logging.warning("⚠️ No se encontraron datos para REGION OCCIDENTE")
            return None
        
        # Excluir columnas no deseadas (case-insensitive)
        columnas_excluir_lower = [col.lower() for col in COLUMNAS_EXCLUIR]
        columnas_a_mantener = [col for col in df.columns if col.lower() not in columnas_excluir_lower]
        df = df[columnas_a_mantener]
        
        logging.info(f"📊 Datos obtenidos: {len(df)} registros de REGION OCCIDENTE ({len(columnas_a_mantener)} columnas)")
        
        return df
        
    except Exception as e:
        logging.error(f"❌ Error obteniendo datos de REGION OCCIDENTE: {e}")
        return None
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass


def clean_excel_illegal_chars(text):
    """
    Limpiar caracteres ilegales que Excel no acepta en celdas
    Excel no permite caracteres de control (0x00-0x1F excepto tab, newline, carriage return)
    """
    if text is None or not isinstance(text, str):
        return text
    
    import re
    # Remover caracteres de control ilegales para Excel (excepto \t, \n, \r)
    illegal_chars = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')
    return illegal_chars.sub('', text)


def format_excel_workbook(filepath: str):
    """
    Aplicar formato profesional al archivo Excel
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = load_workbook(filepath)
    ws = wb.active
    
    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="left", vertical="center")
    cell_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Colores alternados para filas
    row_fill_even = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Aplicar formato a encabezados (fila 1)
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border
    
    # Aplicar formato a datos y filas alternadas
    for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
        fill = row_fill_even if row_num % 2 == 0 else row_fill_odd
        for cell in row:
            cell.alignment = cell_alignment
            cell.border = cell_border
            cell.fill = fill
    
    # Ajustar ancho de columnas automáticamente
    for col_num in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        
        for row_num in range(1, min(ws.max_row + 1, 100)):  # Revisar primeras 100 filas
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value:
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)
        
        # Limitar ancho entre 10 y 50 caracteres
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Congelar primera fila (encabezados)
    ws.freeze_panes = 'A2'
    
    # Agregar filtros automáticos
    ws.auto_filter.ref = ws.dimensions
    
    # Guardar cambios
    wb.save(filepath)
    wb.close()
    
    logging.info("✨ Formato profesional aplicado al Excel")


def create_excel_region_occidente(output_path: str = None) -> Optional[str]:
    """
    Crear archivo Excel con datos filtrados por REGION OCCIDENTE
    """
    try:
        # Obtener datos filtrados (retorna pandas DataFrame)
        df = get_region_occidente_data()
        
        if df is None or df.empty:
            logging.error("❌ No hay datos para crear el Excel de REGION OCCIDENTE")
            return None
        
        # Limpiar caracteres ilegales en columnas de texto
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].apply(clean_excel_illegal_chars)
        
        # Generar nombre de archivo si no se proporciona
        if output_path is None:
            timestamp = now_colombia().strftime('%d-%m-%Y_%H.%M')
            output_path = os.path.join(DOWNLOAD_DIR, f"Backlog_Region_Occidente_{timestamp}.xlsx")
        
        # Asegurar que existe la carpeta de descargas
        ensure_download_folder()
        
        # Escribir Excel usando pandas
        df.to_excel(output_path, index=False, engine='openpyxl')
        
        # Aplicar formato profesional
        format_excel_workbook(output_path)
        
        logging.info(f"✅ Excel creado exitosamente: {output_path}")
        logging.info(f"📊 Total de registros: {len(df)}")
        
        return output_path
        
    except Exception as e:
        logging.error(f"❌ Error creando Excel de REGION OCCIDENTE: {e}")
        return None


def send_email_with_attachment(
    subject: str,
    body: str,
    attachment_path: str = None,
    recipients: list = None
) -> bool:
    """
    Enviar email con archivo adjunto usando Gmail SMTP
    """
    try:
        # Usar configuración por defecto si no se proporcionan destinatarios
        if recipients is None:
            recipients = EMAIL_CONFIG['recipients']
        
        # Crear mensaje
        msg = MIMEMultipart()
        msg['From'] = EMAIL_CONFIG['sender_email']
        msg['To'] = ', '.join(recipients)
        msg['Subject'] = subject
        
        # Agregar cuerpo del mensaje
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        # Agregar archivo adjunto si existe
        if attachment_path and os.path.exists(attachment_path):
            with open(attachment_path, 'rb') as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
            
            encoders.encode_base64(part)
            filename = os.path.basename(attachment_path)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename= {filename}'
            )
            msg.attach(part)
            logging.info(f"📎 Archivo adjunto agregado: {filename}")
        
        # Conectar y enviar
        logging.info(f"📧 Conectando a servidor SMTP: {EMAIL_CONFIG['smtp_server']}:{EMAIL_CONFIG['smtp_port']}")
        
        server = smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port'])
        server.starttls()
        server.login(EMAIL_CONFIG['sender_email'], EMAIL_CONFIG['app_password'])
        
        text = msg.as_string()
        server.sendmail(EMAIL_CONFIG['sender_email'], recipients, text)
        server.quit()
        
        logging.info(f"✅ Email enviado exitosamente a: {', '.join(recipients)}")
        return True
        
    except smtplib.SMTPAuthenticationError as e:
        logging.error(f"❌ Error de autenticación SMTP: {e}")
        return False
    except Exception as e:
        logging.error(f"❌ Error enviando email: {e}")
        return False


def send_region_occidente_report() -> bool:
    """
    Función principal que:
    1. Crea el Excel filtrado por REGION OCCIDENTE
    2. Envía el email con el archivo adjunto
    """
    try:
        logging.info("=" * 60)
        logging.info("📊 INICIANDO REPORTE REGION OCCIDENTE")
        logging.info("=" * 60)
        
        # Crear Excel
        excel_path = create_excel_region_occidente()
        
        if not excel_path:
            logging.error("❌ No se pudo crear el Excel, abortando envío de email")
            return False
        
        # Preparar email
        fecha_actual = now_colombia().strftime('%d/%m/%Y %H:%M')
        subject = f"Backlog Region Occidente - {fecha_actual}"
        body = f"""Hola,

Se adjunta el reporte de Backlog Nacional filtrado por REGION OCCIDENTE.

Fecha de generación: {fecha_actual}

Este es un mensaje automático generado por el sistema de monitoreo.

Saludos,
Bot SharePoint Monitor
"""
        
        # Enviar email
        success = send_email_with_attachment(
            subject=subject,
            body=body,
            attachment_path=excel_path
        )
        
        # Limpiar archivo temporal después de enviar
        if success and os.path.exists(excel_path):
            try:
                os.remove(excel_path)
                logging.info(f"🗑️ Archivo temporal eliminado: {excel_path}")
            except Exception as e:
                logging.warning(f"⚠️ No se pudo eliminar archivo temporal: {e}")
        
        if success:
            logging.info("✅ REPORTE REGION OCCIDENTE ENVIADO EXITOSAMENTE")
        else:
            logging.error("❌ FALLO EN ENVÍO DE REPORTE REGION OCCIDENTE")
        
        logging.info("=" * 60)
        return success
        
    except Exception as e:
        logging.error(f"❌ Error en send_region_occidente_report: {e}")
        return False


# Endpoint Flask para enviar reporte manualmente
@app.route('/send-report-occidente', methods=['POST', 'GET'])
def send_report_occidente_endpoint():
    """Endpoint para enviar reporte de REGION OCCIDENTE por email"""
    try:
        success = send_region_occidente_report()
        
        if success:
            return jsonify({
                "status": "success",
                "message": "Reporte de REGION OCCIDENTE enviado exitosamente",
                "timestamp": now_colombia().isoformat()
            }), 200
        else:
            return jsonify({
                "status": "error",
                "message": "Error enviando reporte de REGION OCCIDENTE",
                "timestamp": now_colombia().isoformat()
            }), 500
            
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e),
            "timestamp": now_colombia().isoformat()
        }), 500


if __name__ == "__main__":
    # Configurar el monitor del sistema
    create_system_monitor()
    
    # Ejecutar con alta disponibilidad
    run_with_high_availability()